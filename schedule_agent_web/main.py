"""
Schedule Agent Web — FastAPI backend.
Exposes VueLogic (scheduling-agent) via chat API using the skill as system prompt.
"""
from pathlib import Path
import os
import json
from datetime import datetime

from dotenv import load_dotenv

# Load .env from project root (folder containing schedule_agent_web) so it works regardless of CWD
_REPO_ROOT = Path(__file__).resolve().parent.parent
load_dotenv(_REPO_ROOT / ".env")
load_dotenv()  # also allow CWD .env

from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response
from pydantic import BaseModel

# Optional: OpenAI and Anthropic (Claude)
try:
    from openai import OpenAI
except ImportError:
    OpenAI = None
try:
    from anthropic import Anthropic
except ImportError:
    Anthropic = None

app = FastAPI(
    title="VueLogic API",
    description="VueLogic — Project Controls AI (CPM, WBS, DCMA 14-Point, Baseline Review, Schedule Intelligence)",
    version="0.3.0-beta",
)

# Seed auth database (clients + default users) on startup
try:
    from schedule_agent_web.auth import ensure_seeded
    ensure_seeded()
except Exception:
    pass

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Paths: assume repo root is parent of schedule_agent_web
REPO_ROOT = Path(__file__).resolve().parent.parent
SKILL_PATH = REPO_ROOT / ".cursor" / "skills" / "scheduling-agent" / "SKILL.md"

# Override with env if set (e.g. in Docker)
if os.environ.get("SCHEDULE_AGENT_SKILL_PATH"):
    SKILL_PATH = Path(os.environ["SCHEDULE_AGENT_SKILL_PATH"])


def get_system_prompt() -> str:
    """Load scheduling-agent skill as system prompt."""
    scope_instruction = (
        "You are VueLogic, a project controls assistant. Your primary expertise is project scheduling "
        "(CPM, WBS, logic, baselines, P6, delays, re-sequencing, critical path, what-if, DCMA 14-Point), "
        "cost controls (EVM, Earned Schedule, EAC), risk management, and MEP/commissioning.\n\n"
        "You also help users with questions about their project data and documents — such as what files "
        "are uploaded, their vectorization status, what content is in their reference library, "
        "and any questions about the documents they have submitted (narratives, XER files, specs, etc.).\n\n"
        "For topics completely unrelated to project controls or the user's project data, "
        "politely redirect to project-related questions.\n\n"
    )
    if not SKILL_PATH.exists():
        return (
            scope_instruction
            + "You help with CPM schedules, WBS, logic, DCMA 14-Point, re-sequencing, and what-if analysis. "
            "Skill file not found; using default behavior."
        )
    return scope_instruction + SKILL_PATH.read_text(encoding="utf-8", errors="replace")


SYSTEM_PROMPT = None


def get_system_prompt_cached() -> str:
    global SYSTEM_PROMPT
    if SYSTEM_PROMPT is None:
        SYSTEM_PROMPT = get_system_prompt()
    return SYSTEM_PROMPT


def _friendly_file_label(filename: str) -> str:
    """Convert a raw filename into a user-friendly label for AI context."""
    name = filename.rsplit(".", 1)[0] if "." in filename else filename
    name = name.replace("_", " ").replace("-", " ").strip()
    ext = filename.rsplit(".", 1)[-1].upper() if "." in filename else ""
    lower = name.lower()
    if any(kw in lower for kw in ["spec", "provision", "technical"]):
        return f"Contract Specifications ({ext})"
    if any(kw in lower for kw in ["narrative", "baseline schedule"]):
        return f"Schedule Narrative ({ext})"
    if any(kw in lower for kw in ["scope", "contract"]):
        return f"Scope Document ({ext})"
    return f"Project Document: {name} ({ext})" if ext else f"Project Document: {name}"


def get_system_prompt_with_context(session_id: str | None = None) -> str:
    """Base skill + lessons learned + trust score (if Redis) + uploaded files (always when session_id set)."""
    base = get_system_prompt_cached()
    parts = [base]
    try:
        from schedule_agent_web.store import get_lessons, get_trust_score, is_persistence_available, get_files, get_file_content
        if is_persistence_available():
            lessons = get_lessons()
            trust = get_trust_score()
            if lessons:
                parts.append("\n\n## Current lessons learned (use when proposing options)\n")
                for i, le in enumerate(lessons[-20:], 1):  # last 20
                    parts.append(f"- [{i}] {le.get('event', '')}: {le.get('lesson', '')}\n")
            parts.append("\n\n## Trust score (HITL)\n")
            parts.append(f"Approvals: {trust.get('approvals', 0)}, Total proposals: {trust.get('total_proposals', 0)}, AI_Agency_Score: {trust.get('ai_agency_score', 0)}. Level 1 (Autonomous) only if score ≥ 0.8; otherwise propose (Level 2/3).\n")
        # Include uploaded files whenever session_id is provided (works with Redis or local file store)
        if session_id:
            files = get_files(session_id)
            user_files = [f for f in files if not f.get("filename", "").startswith("_")]
            if user_files:
                vec_count = sum(1 for f in user_files if f.get("vectorized"))
                parts.append(
                    f"\n\n## Project reference data ({len(user_files)} files, {vec_count} vectorized)\n"
                    "IMPORTANT: The data below comes from the project's reference library. "
                    "NEVER reveal internal file names, storage paths, file prefixes, or library structure to the user. "
                    "When referencing data, say 'the project schedule' or 'the contract specifications' — "
                    "never mention raw filenames or storage paths to the user.\n"
                    "When the user asks about files, vectorization, or their library, answer using the file list below.\n"
                )
                context_budget = 120000
                context_used = 0
                import re as _re
                for f in user_files:
                    filename = f.get("filename", "")
                    content = get_file_content(session_id, filename)
                    if not content:
                        continue
                    is_xer = filename.lower().endswith(".xer")
                    is_pdf = filename.lower().endswith(".pdf")
                    is_base64 = bool(_re.match(r'^[A-Za-z0-9+/\r\n]{100,}=*$', content[:200].replace('\n','').replace('\r','')))
                    if is_pdf and is_base64:
                        label = _friendly_file_label(filename)
                        try:
                            from schedule_agent_web.vector_store import _extract_text_from_pdf_bytes
                            pdf_text = _extract_text_from_pdf_bytes(content)
                            if pdf_text and pdf_text != content and len(pdf_text) > 50:
                                budget_left = min(6000, context_budget - context_used)
                                if budget_left > 500:
                                    preview = pdf_text[:budget_left]
                                    parts.append(f"\n### {label} (full document available — RAG retrieval provides deeper content on demand)\n```\n{preview}\n```\n")
                                    context_used += len(preview)
                                else:
                                    parts.append(f"\n### {label} (PDF document available in library)\n")
                            else:
                                parts.append(f"\n### {label} (PDF document available in library)\n")
                        except Exception:
                            parts.append(f"\n### {label} (PDF document available in library)\n")
                        continue
                    if is_xer:
                        label = _friendly_file_label(filename)
                        version_hint = ""
                        fn_lower = filename.lower()
                        if "baseline_v1" in fn_lower or "_bas." in fn_lower:
                            version_hint = " — Baseline V1"
                        elif "baseline_v2" in fn_lower or "_bas-r1" in fn_lower or "r1." in fn_lower:
                            version_hint = " — Baseline V2"
                        elif "update_v" in fn_lower:
                            import re as _re2
                            m = _re2.search(r'update_v(\d+)', fn_lower)
                            version_hint = f" — Update V{m.group(1)}" if m else " — Update"
                        content = _decode_xer_content(content)
                        xer_summary = _summarize_xer_for_review(content)
                        budget_per_xer = min(55000, context_budget - context_used)
                        if budget_per_xer < 500:
                            parts.append(f"\n### Project Schedule (P6 XER{version_hint}) (budget exceeded — summary not included)\n")
                            continue
                        preview = xer_summary[:budget_per_xer]
                        parts.append(f"\n### Project Schedule (P6 XER{version_hint})\n```\n{preview}\n```\n")
                        context_used += len(preview)
                    else:
                        label = _friendly_file_label(filename)
                        budget_left = min(50000, context_budget - context_used)
                        if budget_left < 500:
                            continue
                        preview = content[:budget_left] if len(content) > budget_left else content
                        parts.append(f"\n### {label}\n```\n{preview}\n```\n")
                        context_used += len(preview)
        # Inject philosophy / governance settings so the agent applies them
        if session_id:
            try:
                from schedule_agent_web.store import get_philosophy
                phil = get_philosophy(session_id)
                gov = phil.get("governance", {})
                bas = phil.get("basis", {})
                pb = phil.get("playbook", {})
                philo_block = "\n\n## Active Schedule Philosophy (this project)\n"
                philo_block += "IMPORTANT: Apply these rules in ALL schedule reviews, baseline analyses, and chat responses about schedules.\n\n"
                philo_block += "### Governance Thresholds\n"
                std_map = {"dcma": "DCMA 14-Point", "aace": "AACE International", "custom": "Owner Custom"}
                tone_map = {"exec": "Executive", "tech": "Technical", "aggr": "Aggressive", "cons": "Conservative"}
                philo_block += f"- Compliance standard: {std_map.get(gov.get('complianceStandard', 'dcma'), gov.get('complianceStandard', 'dcma'))}\n"
                philo_block += f"- Missing logic tolerance: {gov.get('missingLogicTolerance', 5)}%\n"
                philo_block += f"- High float threshold: {gov.get('highFloatDays', 44)} days\n"
                philo_block += f"- High duration threshold: {gov.get('highDurationDays', 20)} days\n"
                philo_block += f"- Negative lag tolerance: {gov.get('negativeLagTolerance', 0)}\n"
                philo_block += f"- Hard constraint tolerance: {gov.get('hardConstraintTolerance', 5)}%\n"
                rules = gov.get("rules", {})
                active_rules = []
                if rules.get("leadRestriction"): active_rules.append("Negative lags (leads) are FORBIDDEN")
                if rules.get("sfBan"): active_rules.append("Start-to-Finish relationships are BANNED")
                if rules.get("hardConstraintAudit"): active_rules.append("Flag all hard constraints as deficiencies")
                if rules.get("calendarCheck"): active_rules.append("Every activity MUST have an assigned calendar")
                if rules.get("fsPreferred"): active_rules.append("Flag when FS relationships < 80%")
                if active_rules:
                    philo_block += "- Active rules: " + "; ".join(active_rules) + "\n"
                philo_block += f"- Narrative tone: {tone_map.get(gov.get('narrativeTone', 'exec'), gov.get('narrativeTone', 'exec'))} — write all review comments and analysis in this tone.\n"
                philo_block += "\n### Schedule Basis Context\n"
                philo_block += f"- Delivery method: {bas.get('deliveryMethod', 'Design-Build')}\n"
                if bas.get("ntpDate"): philo_block += f"- NTP date: {bas['ntpDate']}\n"
                if bas.get("completionDate"): philo_block += f"- Completion date: {bas['completionDate']}\n"
                philo_block += f"- WBS breakdown: {bas.get('wbsDriver', 'area')} driven, {bas.get('wbsLevels', 4)} levels\n"
                if bas.get("wbsRationale"): philo_block += f"- WBS rationale: {bas['wbsRationale']}\n"
                philo_block += f"- Calendar: {bas.get('workWeek', '5-day')} week, {bas.get('hoursPerDay', 8)} hrs/day, {bas.get('shiftsPerDay', 1)} shift(s)\n"
                if bas.get("weatherNotes"): philo_block += f"- Weather/seasonal: {bas['weatherNotes']}\n"
                lr = bas.get("logicRationale", {})
                if lr.get("fsPref"): philo_block += f"- FS-preferred rationale: {lr['fsPref']}\n"
                if lr.get("zeroLead"): philo_block += f"- Zero-lead rationale: {lr['zeroLead']}\n"
                if lr.get("floatThreshold"): philo_block += f"- Float threshold rationale: {lr['floatThreshold']}\n"
                philo_block += f"- Cost loading: {bas.get('costLoadingLevel', 'work-package')} level\n"
                if bas.get("costLoadingNotes"): philo_block += f"- Cost loading notes: {bas['costLoadingNotes']}\n"
                philo_block += "\n### Baseline Review Playbook\n"
                std_rev = {"approve-noted": "Approve as Noted (preferred)", "approve-full": "Full Approval Required", "reject-resubmit": "Reject & Resubmit"}
                philo_block += f"- Review standard: {std_rev.get(pb.get('reviewStandard', 'approve-noted'), pb.get('reviewStandard', 'approve-noted'))}\n"
                philo_block += f"- Max deficiencies before rejection: {pb.get('maxDefectsBeforeReject', 20)}\n"
                philo_block += f"- Missing scope rejection threshold: {pb.get('missingScopeThreshold', 10)}%\n"
                cl = pb.get("checklist", {})
                required_checks = [k for k, v in cl.items() if v]
                if required_checks:
                    philo_block += "- Required review checks: " + ", ".join(k.replace("_", " ").title() for k in required_checks) + "\n"
                oa = pb.get("ownerActivities", [])
                if oa:
                    req_oa = [a for a in oa if a.get("required")]
                    if req_oa:
                        philo_block += "- Required owner activities in schedule: " + ", ".join(f"{a['name']} ({a['duration']}d)" for a in req_oa) + "\n"
                parts.append(philo_block)
            except Exception:
                pass
        return "".join(parts)
    except Exception:
        return base


class ChatRequest(BaseModel):
    message: str
    history: list[dict[str, str]] = []
    session_id: str | None = None


class ChatResponse(BaseModel):
    reply: str
    error: str | None = None
    sources: list[dict[str, str]] | None = None  # Stage 4: cite source doc [{ filename, text_preview }]


@app.get("/health")
def health():
    return {"status": "ok", "agent": "schedule-agent"}


# ── Authentication endpoints ────────────────────────────────────────────────

class RegisterRequest(BaseModel):
    first_name: str
    last_name: str
    email: str
    company: str
    password: str


class LoginRequest(BaseModel):
    email: str
    password: str


@app.post("/api/auth/register")
def api_register(req: RegisterRequest):
    from schedule_agent_web.auth import register_user
    result = register_user(req.first_name, req.last_name, req.email, req.company, req.password)
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


@app.post("/api/auth/login")
def api_login(req: LoginRequest):
    from schedule_agent_web.auth import login_user
    result = login_user(req.email, req.password)
    if not result["ok"]:
        raise HTTPException(status_code=401, detail=result["error"])
    return result


@app.get("/api/auth/me")
def api_auth_me(authorization: str = Header(default="")):
    """Verify the current token and return the user profile."""
    from schedule_agent_web.auth import verify_token
    token = authorization.replace("Bearer ", "") if authorization.startswith("Bearer ") else authorization
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    user = verify_token(token)
    if user is None:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    return {"ok": True, "user": user}


@app.get("/api/auth/projects")
def api_auth_projects(authorization: str = Header(default="")):
    """Return the list of projects available for the authenticated user's company."""
    from schedule_agent_web.auth import get_projects_for_token
    token = authorization.replace("Bearer ", "") if authorization.startswith("Bearer ") else authorization
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    projects = get_projects_for_token(token)
    return {"ok": True, "projects": projects}


class CreateProjectRequest(BaseModel):
    project_id: str
    name: str = ""


@app.post("/api/auth/projects")
def api_create_project(req: CreateProjectRequest, authorization: str = Header(default="")):
    """Create a new project under the authenticated user's company."""
    from schedule_agent_web.auth import verify_token, create_project
    token = authorization.replace("Bearer ", "") if authorization.startswith("Bearer ") else authorization
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    user = verify_token(token)
    if user is None:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    result = create_project(req.project_id, req.name, user.get("client_id"))
    if not result["ok"]:
        raise HTTPException(status_code=400, detail=result["error"])
    return result


def _get_openai_key():
    raw = os.environ.get("OPENAI_API_KEY") or os.environ.get("OPENAI_API_KEY_FILE")
    if raw and Path(raw).is_file():
        return Path(raw).read_text().strip()
    return raw or ""


def _get_anthropic_key():
    raw = os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_API_KEY_FILE")
    if raw and Path(raw).is_file():
        return Path(raw).read_text().strip()
    return raw or ""


def _use_claude():
    """Use Claude if Anthropic key is set (overrides OpenAI when both set)."""
    return bool(_get_anthropic_key() and len(_get_anthropic_key()) > 10 and Anthropic is not None)


def get_status_dict():
    """Debug: confirm API key and skill (callable from api/status.py)."""
    openai_key = _get_openai_key()
    anthropic_key = _get_anthropic_key()
    has_openai = bool(openai_key and len(openai_key) > 10)
    has_anthropic = bool(anthropic_key and len(anthropic_key) > 10)
    skill_ok = SKILL_PATH.exists()
    provider = "claude" if (has_anthropic and Anthropic) else ("openai" if (has_openai and OpenAI) else None)
    return {
        "status": "ok",
        "has_api_key": has_openai or has_anthropic,
        "has_anthropic_key": has_anthropic,
        "has_openai_key": has_openai,
        "provider": provider,
        "skill_loaded": skill_ok,
        "openai_installed": OpenAI is not None,
        "anthropic_installed": Anthropic is not None,
    }


@app.get("/api/status")
def api_status():
    d = get_status_dict()
    try:
        from schedule_agent_web.store import is_persistence_available
        d["persistence_available"] = is_persistence_available()
    except Exception:
        d["persistence_available"] = False
    try:
        from schedule_agent_web.vector_store import is_qdrant_available, is_graphrag_available
        d["qdrant_available"] = is_qdrant_available()
        d["graphrag_available"] = is_graphrag_available()
    except Exception:
        d["qdrant_available"] = False
        d["graphrag_available"] = False
    return d


@app.get("/api/greeting")
def api_greeting(session_id: str = ""):
    """Generate a dynamic, project-aware greeting for the dashboard chat."""
    if not session_id:
        return {"greeting": "How can I help you with project controls today?"}
    import random
    from datetime import datetime, timedelta

    proj_id = session_id
    now = datetime.now()
    hour = now.hour
    time_greeting = "Good morning" if hour < 12 else ("Good afternoon" if hour < 17 else "Good evening")

    observations: list[str] = []
    nudges: list[str] = []

    try:
        from schedule_agent_web.store import get_files
        from schedule_agent_web.baseline import list_submissions, list_reviews
        from schedule_agent_web.store import get_pending_digest, get_approved_log

        files = get_files(session_id)
        user_files = [f for f in files if not f.get("filename", "").startswith("_")]
        vec_count = sum(1 for f in user_files if f.get("vectorized"))

        if user_files:
            observations.append(f"Your library has {len(user_files)} document{'s' if len(user_files) != 1 else ''} ({vec_count} vectorized).")

        baseline_subs = list_submissions(session_id, "baseline")
        update_subs = list_submissions(session_id, "update")
        reviews = list_reviews(session_id)
        review_versions = {r.get("version") for r in reviews}

        if not baseline_subs:
            nudges.append("No baseline schedule has been submitted yet. Would you like to start a Contractor Baseline Review?")
        elif baseline_subs:
            latest_ver = max(s.get("version", 0) for s in baseline_subs)
            if latest_ver not in review_versions:
                nudges.append(f"Baseline V{latest_ver} has been submitted but not yet reviewed. Ready to run the AI-powered review?")
            else:
                review = next((r for r in reviews if r.get("version") == latest_ver), None)
                if review:
                    count = review.get("comment_count", 0)
                    has_response = any(
                        "resp" in f.get("filename", "").lower()
                        for f in user_files
                        if f.get("filename", "").startswith(f"baseline_v{latest_ver}_")
                    )
                    if has_response:
                        observations.append(f"Baseline V{latest_ver} review ({count} comments) has a contractor response on file.")
                    else:
                        nudges.append(f"V{latest_ver} review generated {count} comments. Have you received the contractor's response yet?")

            if update_subs:
                latest_upd = max(s.get("version", 0) for s in update_subs)
                observations.append(f"Schedule Update V{latest_upd} is on record.")

        pending_digest = get_pending_digest(session_id)
        if pending_digest and len(pending_digest) >= 3:
            nudges.append(f"You have {len(pending_digest)} conversation{'s' if len(pending_digest) != 1 else ''} pending review for vectorization.")

        approved = get_approved_log(session_id)
        not_vec = [i for i in approved if not i.get("vectorized", False)]
        if not_vec:
            nudges.append(f"{len(not_vec)} approved conversation{'s' if len(not_vec) != 1 else ''} still need vectorization.")

        if not user_files:
            nudges.append("Your project library is empty. Upload contract specs, schedule narratives, or XER files to get started.")

    except Exception:
        pass

    meeting_prompts = [
        "Did you attend any recent meetings that I should know about? Just say 'FYI' or 'Remember this' to share.",
        "Anything from site or a coordination meeting you'd like me to capture? Start with 'FYI' and I'll log it.",
        "If you've picked up new information from the project team, share it with me so I can stay current.",
    ]

    parts = [f"{time_greeting} \u2014 you're working on **{proj_id}**."]
    if observations:
        parts.append(" ".join(observations))
    if nudges:
        parts.append(nudges[0])
    elif not observations:
        parts.append("How can I help you with project controls today?")
    else:
        parts.append(random.choice(meeting_prompts))

    greeting = " ".join(parts)
    return {"greeting": greeting, "project_id": proj_id}


@app.get("/api/philosophy")
def api_philosophy_get(session_id: str = ""):
    """Return philosophy settings for this project."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import get_philosophy
    return get_philosophy(session_id)


@app.put("/api/philosophy")
def api_philosophy_put(body: dict):
    """Save philosophy settings for this project."""
    session_id = body.get("session_id", "")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import save_philosophy, get_philosophy
    current = get_philosophy(session_id)
    for section in ("governance", "basis", "playbook"):
        if section in body:
            if isinstance(current.get(section), dict) and isinstance(body[section], dict):
                current[section].update(body[section])
            else:
                current[section] = body[section]
    ok = save_philosophy(session_id, current)
    if not ok:
        raise HTTPException(status_code=500, detail="Failed to save")
    return {"ok": True, "philosophy": current}


@app.get("/api/schedule/intelligence")
def api_schedule_intelligence(session_id: str = ""):
    """Analyze uploaded XER against governance thresholds and return DCMA scorecard + metrics."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import get_files, get_file_content, get_philosophy

    files = get_files(session_id)
    xer_file = None
    for f in files:
        fn = f.get("filename", "")
        if fn.lower().endswith(".xer") and not fn.startswith("_"):
            xer_file = fn
    if not xer_file:
        from schedule_agent_web.baseline import list_submissions
        subs = list_submissions(session_id, submission_type="baseline")
        if subs:
            latest = subs[-1]
            xer_file = f"baseline_v{latest['version']}_{latest['xer_filename']}"
    if not xer_file:
        return {"error": "no_xer", "message": "Upload an XER schedule to see intelligence data."}

    raw = get_file_content(session_id, xer_file) or ""
    xer_text = _decode_xer_content(raw)
    if not xer_text or len(xer_text) < 50:
        return {"error": "empty_xer", "message": "XER file is empty or unreadable."}

    tables = _parse_xer_tables(xer_text)
    phil = get_philosophy(session_id)
    gov = phil.get("governance", {})

    tasks = tables.get("TASK", [])
    preds = tables.get("TASKPRED", [])
    wbs = tables.get("PROJWBS", [])
    cals = tables.get("CALENDAR", [])
    rsrc = tables.get("TASKRSRC", [])

    total = len(tasks)
    if total == 0:
        return {"error": "no_tasks", "message": "XER contains no activities."}

    # Build predecessor/successor maps
    task_ids = {t.get("task_id") for t in tasks}
    has_pred = set()
    has_succ = set()
    rel_types = {"FS": 0, "SS": 0, "FF": 0, "SF": 0}
    neg_lag_count = 0
    cross_wbs_pref = 0
    task_wbs = {t.get("task_id"): t.get("wbs_id", "") for t in tasks}

    for p in preds:
        tid = p.get("task_id", "")
        pid = p.get("pred_task_id", "")
        has_pred.add(tid)
        has_succ.add(pid)
        pt = (p.get("pred_type") or "FS").upper().strip()
        if pt in rel_types:
            rel_types[pt] += 1
        lag = p.get("lag_hr_cnt", "0")
        try:
            lag_val = float(lag)
            if lag_val < 0:
                neg_lag_count += 1
        except (ValueError, TypeError):
            pass
        if task_wbs.get(tid, "") != task_wbs.get(pid, ""):
            cross_wbs_pref += 1

    total_rels = sum(rel_types.values())
    fs_pct = round(rel_types["FS"] / total_rels * 100, 1) if total_rels else 0

    # Missing logic (no pred or no succ) — exclude milestones
    task_type_map = {t.get("task_id"): (t.get("task_type") or "").upper() for t in tasks}
    missing_logic = 0
    for t in tasks:
        tid = t.get("task_id", "")
        tt = (t.get("task_type") or "").upper()
        if tt in ("TT_Mile", "TT_FinMile", "TT_WBS", "TT_LOE"):
            continue
        if tid not in has_pred or tid not in has_succ:
            missing_logic += 1
    missing_logic_pct = round(missing_logic / total * 100, 1)

    # High float
    high_float_threshold = gov.get("highFloatDays", 44)
    high_float_count = 0
    float_values = []
    for t in tasks:
        tt = (t.get("task_type") or "").upper()
        if tt in ("TT_Mile", "TT_FinMile", "TT_WBS", "TT_LOE"):
            continue
        try:
            tf_hrs = float(t.get("total_float_hr_cnt", 0))
            tf_days = tf_hrs / 8
            float_values.append(tf_days)
            if tf_days > high_float_threshold:
                high_float_count += 1
        except (ValueError, TypeError):
            pass
    high_float_pct = round(high_float_count / total * 100, 1) if total else 0

    # Negative float
    neg_float = sum(1 for f in float_values if f < 0)
    neg_float_pct = round(neg_float / total * 100, 1) if total else 0

    # High duration
    high_dur_threshold = gov.get("highDurationDays", 20)
    high_dur_count = 0
    durations = []
    for t in tasks:
        tt = (t.get("task_type") or "").upper()
        if tt in ("TT_Mile", "TT_FinMile", "TT_WBS", "TT_LOE"):
            continue
        try:
            dur_hrs = float(t.get("target_drtn_hr_cnt", 0))
            dur_days = dur_hrs / 8
            durations.append(dur_days)
            if dur_days > high_dur_threshold:
                high_dur_count += 1
        except (ValueError, TypeError):
            pass
    high_dur_pct = round(high_dur_count / total * 100, 1) if total else 0

    # Hard constraints
    constraint_types = ("CS_MSO", "CS_MSOB", "CS_MFO", "CS_MFOB")
    hard_const_count = sum(1 for t in tasks if (t.get("cstr_type") or "").upper() in constraint_types or (t.get("cstr_type2") or "").upper() in constraint_types)
    hard_const_pct = round(hard_const_count / total * 100, 1) if total else 0

    # Relationship type distribution
    sf_count = rel_types.get("SF", 0)

    # Resource loading
    tasks_with_rsrc = set(r.get("task_id") for r in rsrc)
    rsrc_loaded_pct = round(len(tasks_with_rsrc) / total * 100, 1) if total else 0

    # Calendar coverage
    tasks_with_cal = sum(1 for t in tasks if t.get("clndr_id"))
    cal_pct = round(tasks_with_cal / total * 100, 1) if total else 0

    # Thresholds from governance
    ml_threshold = gov.get("missingLogicTolerance", 5)
    hc_threshold = gov.get("hardConstraintTolerance", 5)
    nl_threshold = gov.get("negativeLagTolerance", 0)

    def _status(actual, threshold, higher_is_bad=True):
        if higher_is_bad:
            return "pass" if actual <= threshold else ("warn" if actual <= threshold * 1.5 else "fail")
        return "pass" if actual >= threshold else ("warn" if actual >= threshold * 0.8 else "fail")

    scorecard = [
        {"id": 1, "criterion": "Missing Logic", "target": f"≤ {ml_threshold}%", "actual": f"{missing_logic_pct}%", "status": _status(missing_logic_pct, ml_threshold)},
        {"id": 2, "criterion": "Missing Predecessors", "target": "0", "actual": str(sum(1 for t in tasks if t.get("task_id") not in has_pred and (t.get("task_type") or "").upper() not in ("TT_Mile","TT_FinMile","TT_WBS","TT_LOE"))), "status": _status(sum(1 for t in tasks if t.get("task_id") not in has_pred and (t.get("task_type") or "").upper() not in ("TT_Mile","TT_FinMile","TT_WBS","TT_LOE")), 0)},
        {"id": 3, "criterion": "Missing Successors", "target": "0", "actual": str(sum(1 for t in tasks if t.get("task_id") not in has_succ and (t.get("task_type") or "").upper() not in ("TT_Mile","TT_FinMile","TT_WBS","TT_LOE"))), "status": _status(sum(1 for t in tasks if t.get("task_id") not in has_succ and (t.get("task_type") or "").upper() not in ("TT_Mile","TT_FinMile","TT_WBS","TT_LOE")), 0)},
        {"id": 4, "criterion": "High Float", "target": f"≤ {high_float_threshold}d", "actual": f"{high_float_pct}%", "status": _status(high_float_pct, 5)},
        {"id": 5, "criterion": "Negative Float", "target": "0%", "actual": f"{neg_float_pct}%", "status": _status(neg_float_pct, 0)},
        {"id": 6, "criterion": "High Duration", "target": f"≤ {high_dur_threshold}d", "actual": f"{high_dur_pct}%", "status": _status(high_dur_pct, 5)},
        {"id": 7, "criterion": "Negative Lags (Leads)", "target": str(nl_threshold), "actual": str(neg_lag_count), "status": _status(neg_lag_count, nl_threshold)},
        {"id": 8, "criterion": "SF Relationships", "target": "0", "actual": str(sf_count), "status": _status(sf_count, 0)},
        {"id": 9, "criterion": "Hard Constraints", "target": f"≤ {hc_threshold}%", "actual": f"{hard_const_pct}%", "status": _status(hard_const_pct, hc_threshold)},
        {"id": 10, "criterion": "FS Relationship %", "target": "≥ 80%", "actual": f"{fs_pct}%", "status": _status(fs_pct, 80, higher_is_bad=False)},
        {"id": 11, "criterion": "Resource Loaded", "target": "> 0%", "actual": f"{rsrc_loaded_pct}%", "status": _status(rsrc_loaded_pct, 1, higher_is_bad=False)},
        {"id": 12, "criterion": "Calendar Assigned", "target": "100%", "actual": f"{cal_pct}%", "status": _status(cal_pct, 95, higher_is_bad=False)},
    ]

    pass_count = sum(1 for s in scorecard if s["status"] == "pass")
    warn_count = sum(1 for s in scorecard if s["status"] == "warn")
    fail_count = sum(1 for s in scorecard if s["status"] == "fail")

    # Critical path top 10 (lowest float)
    work_tasks = [t for t in tasks if (t.get("task_type") or "").upper() not in ("TT_Mile", "TT_FinMile", "TT_WBS", "TT_LOE")]
    def _float_val(t):
        try: return float(t.get("total_float_hr_cnt", 999999))
        except: return 999999
    cp_top = sorted(work_tasks, key=_float_val)[:10]
    critical_path = []
    for t in cp_top:
        try:
            fv = round(float(t.get("total_float_hr_cnt", 0)) / 8, 1)
        except:
            fv = 0
        try:
            dv = round(float(t.get("target_drtn_hr_cnt", 0)) / 8, 1)
        except:
            dv = 0
        critical_path.append({
            "task_code": t.get("task_code", ""),
            "task_name": t.get("task_name", ""),
            "duration": dv,
            "total_float": fv,
            "status": t.get("status_code", ""),
        })

    return {
        "scorecard": scorecard,
        "metrics": {
            "totalActivities": total,
            "totalRelationships": total_rels,
            "passCount": pass_count,
            "warnCount": warn_count,
            "failCount": fail_count,
            "fsPercent": fs_pct,
            "ssPercent": round(rel_types["SS"] / total_rels * 100, 1) if total_rels else 0,
            "ffPercent": round(rel_types["FF"] / total_rels * 100, 1) if total_rels else 0,
            "sfPercent": round(rel_types["SF"] / total_rels * 100, 1) if total_rels else 0,
            "crossWbsTies": cross_wbs_pref,
            "avgFloat": round(sum(float_values) / len(float_values), 1) if float_values else 0,
            "avgDuration": round(sum(durations) / len(durations), 1) if durations else 0,
        },
        "criticalPath": critical_path,
        "xer_filename": xer_file,
    }


@app.get("/api/conversation")
def api_get_conversation(session_id: str = ""):
    """Return stored conversation for this session_id (persistence)."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.store import get_conversation as store_get_conv
        return store_get_conv(session_id)
    except Exception:
        return []


@app.get("/api/conversation/export")
def api_export_conversation_txt(session_id: str = ""):
    """Export chat history as a downloadable .txt file. Optional: save chat history to file."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    try:
        from schedule_agent_web.store import get_conversation
        conv = get_conversation(session_id)
    except Exception:
        conv = []
    from datetime import datetime
    lines = ["VueLogic Chat History", "=" * 40, ""]
    for m in conv:
        role = (m.get("role") or "unknown").capitalize()
        content = (m.get("content") or "").strip()
        created = m.get("created_at")
        if created:
            try:
                dt = datetime.fromisoformat(created.replace("Z", "+00:00"))
                ts = dt.strftime("%Y-%m-%d %H:%M:%S")
            except Exception:
                ts = created
            lines.append(f"[{ts}] {role}:")
        else:
            lines.append(f"{role}:")
        lines.append(content)
        lines.append("")
    body = "\n".join(lines).strip() or "No messages yet."
    filename = f"chat-history-{datetime.utcnow().strftime('%Y-%m-%d')}.txt"
    return Response(
        content=body.encode("utf-8"),
        media_type="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- Generated documents (schedule basis, etc.) — persistent download ---
class SaveDocumentRequest(BaseModel):
    session_id: str
    title: str
    content: str = ""


@app.post("/api/documents")
def api_save_document(request: SaveDocumentRequest):
    """Save a generated document (e.g. from chat) for persistent download."""
    if not request.session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    if not (request.title or "").strip():
        raise HTTPException(status_code=400, detail="title required")
    try:
        from schedule_agent_web.store import save_generated_document
        entry = save_generated_document(request.session_id, request.title.strip(), request.content)
        if not entry:
            raise HTTPException(status_code=500, detail="Failed to save document")
        return entry
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/documents")
def api_list_documents(session_id: str = ""):
    """Return list of saved documents for this session (id, title, created_at)."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.store import get_generated_documents
        return get_generated_documents(session_id)
    except Exception:
        return []


@app.get("/api/documents/search")
def api_documents_search(q: str = "", session_id: str = "", top_k: int = 10, group: str = ""):
    """Semantic search over ingested/uploaded documents (Qdrant). Optional group= filters to contract_specs or sample_schedule."""
    if not (q or "").strip():
        return {"hits": [], "sources": []}
    try:
        from schedule_agent_web.vector_store import search, is_qdrant_available
        if not is_qdrant_available():
            return {"hits": [], "sources": []}
        ingestion_group = (group or "").strip() or None
        hits = search(session_id or None, q.strip(), top_k=max(1, min(top_k, 20)), ingestion_group=ingestion_group)
        sources = [{"filename": h.get("filename", ""), "text_preview": (h.get("text") or "")[:300], "ingestion_group": h.get("ingestion_group")} for h in hits]
        return {"hits": hits, "sources": sources}
    except Exception:
        return {"hits": [], "sources": []}


@app.get("/api/documents/{doc_id}")
def api_download_document(doc_id: str, session_id: str = "", title: str = ""):
    """Return document content for download. Content-Disposition: attachment. Optional title= for filename."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    try:
        from schedule_agent_web.store import get_generated_document_content
        content = get_generated_document_content(session_id, doc_id)
        if content is None:
            raise HTTPException(status_code=404, detail="Document not found")
        import re
        safe = re.sub(r"[^\w\-.\s]", "_", (title or doc_id).strip())[:80] or "document"
        safe = safe.replace(" ", "_")
        filename = f"{safe}.md"
        return Response(
            content=content.encode("utf-8"),
            media_type="text/markdown; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/lessons")
def api_get_lessons():
    """Return lessons learned list."""
    try:
        from schedule_agent_web.store import get_lessons
        return get_lessons()
    except Exception:
        return []


class LessonEntry(BaseModel):
    event: str = ""
    what_happened: str = ""
    outcome: str = ""
    lesson: str = ""
    recommendation: str = ""


@app.post("/api/lessons")
def api_append_lesson(entry: LessonEntry):
    """Append one lesson learned."""
    try:
        from schedule_agent_web.store import append_lesson
        append_lesson(entry.model_dump())
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/trust_score")
def api_get_trust_score():
    """Return trust score (approvals, total_proposals, ai_agency_score)."""
    try:
        from schedule_agent_web.store import get_trust_score
        return get_trust_score()
    except Exception:
        return {"approvals": 0, "total_proposals": 0, "ai_agency_score": 0.0}


class TrustScoreUpdate(BaseModel):
    approved: bool


@app.post("/api/trust_score")
def api_record_proposal(update: TrustScoreUpdate):
    """Record one proposal outcome (e.g. user approved or declined)."""
    try:
        from schedule_agent_web.store import record_proposal, get_trust_score
        record_proposal(update.approved)
        return get_trust_score()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/files")
def api_get_files(session_id: str = ""):
    """Return list of uploaded files for this session (excludes internal/library files)."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.store import get_files
        return [f for f in get_files(session_id) if not f.get("filename", "").startswith("_")]
    except Exception:
        return []


# --- Project Library: categorized file management with vectorization tracking ---

def _library_display_name(filename: str) -> str | None:
    """Convert versioned baseline/update filenames to user-friendly display names.
    Returns None for truly internal files that should stay hidden."""
    import re
    m = re.match(r"^(?:baseline|update)_v\d+_(?:resp_)?(.+)$", filename)
    if m:
        return m.group(1)
    return None


def _library_auto_category(filename: str) -> str | None:
    """Auto-assign category for baseline/update-related files. Returns None if not applicable."""
    import re
    m = re.match(r"^(baseline|update)_v\d+_", filename)
    if not m:
        return None
    kind = m.group(1)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if "resp_" in filename:
        return "Contractor Response"
    if kind == "update":
        if ext == "xer":
            return "Schedule Update (XER)"
        if ext == "pdf":
            return "Update Narrative"
        return "Schedule Update"
    if ext == "xer":
        return "Baseline Schedule (XER)"
    if ext == "pdf":
        return "Baseline Narrative"
    return "Baseline Submission"


# --- Conversation Digest (PCM review & vectorization governance) ---

@app.get("/api/conversation/digest")
def api_digest_get(session_id: str = ""):
    """Get pending digest items for PCM review."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import get_pending_digest
    return {"ok": True, "items": get_pending_digest(session_id)}


def _guess_conversation_category(text: str) -> str:
    """Keyword-based fallback when LLM is unavailable."""
    t = text.lower()
    if any(w in t for w in ["baseline", "xer", "p6", "activity", "activities", "wbs", "logic"]):
        return "Baseline Review"
    if any(w in t for w in ["schedule", "critical path", "float", "delay", "duration", "gantt"]):
        return "Schedule Analysis"
    if any(w in t for w in ["cost", "evm", "earned value", "budget", "eac", "bac", "variance"]):
        return "Cost & EVM"
    if any(w in t for w in ["risk", "qsra", "qcra", "mitigation", "monte carlo"]):
        return "Risk"
    if any(w in t for w in ["file", "document", "vectorize", "library", "upload", "pdf", "narrative"]):
        return "Document Query"
    if any(w in t for w in ["system", "login", "password", "admin", "setting"]):
        return "System Query"
    return "General"


@app.post("/api/conversation/digest/generate")
def api_digest_generate(session_id: str = "", body: dict = {}):
    """Generate a digest from recent unreviewed conversations using AI summarization."""
    if not session_id:
        session_id = body.get("session_id", "")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import (
        get_conversation, get_pending_digest, save_pending_digest,
    )
    conv = get_conversation(session_id)
    if not conv:
        return {"ok": True, "generated": 0, "message": "No conversations to digest."}

    existing = get_pending_digest(session_id)
    existing_ids = {i.get("id") for i in existing}

    last_reviewed_at = body.get("since")

    turns: list[dict] = []
    for i in range(0, len(conv) - 1, 2):
        user_msg = conv[i] if conv[i].get("role") == "user" else None
        asst_msg = conv[i + 1] if i + 1 < len(conv) and conv[i + 1].get("role") == "assistant" else None
        if not user_msg:
            continue
        ts = user_msg.get("created_at", "")
        if last_reviewed_at and ts <= last_reviewed_at:
            continue
        content = user_msg.get("content", "")
        priority = "normal"
        cl = content.strip().lower()
        if cl.startswith("fyi") or cl.startswith("remember this"):
            priority = "fyi"

        turn_id = f"turn-{ts}-{hash(content) & 0xFFFFFFFF:08x}"
        if turn_id in existing_ids:
            continue

        turns.append({
            "id": turn_id,
            "timestamp": ts,
            "user_message": content,
            "assistant_response": asst_msg.get("content", "") if asst_msg else "",
            "priority": priority,
            "status": "pending",
        })

    if not turns:
        return {"ok": True, "generated": 0, "message": "No new conversations to digest."}

    if not get_status_dict().get("has_api_key"):
        for t in turns:
            t["topic"] = t["user_message"][:80] + ("..." if len(t["user_message"]) > 80 else "")
            t["summary"] = t["user_message"][:200]
            t["category"] = _guess_conversation_category(t["user_message"])
        existing.extend(turns)
        save_pending_digest(session_id, existing)
        return {"ok": True, "generated": len(turns), "message": f"Generated {len(turns)} digest items (no AI — used message preview)."}

    batch_text = ""
    for i, t in enumerate(turns):
        batch_text += f"\n--- Turn {i+1} (priority: {t['priority']}) ---\n"
        batch_text += f"User: {t['user_message'][:500]}\n"
        batch_text += f"Assistant: {t['assistant_response'][:500]}\n"

    system = (
        "You are a digest summarizer for a project controls system. Given a list of chat turns, produce a JSON array. "
        "Each element: {\"index\": <0-based>, \"topic\": \"<short topic title max 60 chars>\", "
        "\"summary\": \"<1-2 sentence summary of the exchange>\", "
        "\"category\": \"<one of: Schedule Analysis, Cost & EVM, Risk, Baseline Review, Document Query, System Query, General>\" }. "
        "Pick the single best-fit category. Return ONLY the JSON array, no markdown."
    )
    messages = [{"role": "user", "content": f"Summarize these chat turns:\n{batch_text}"}]
    reply, err = _call_llm(system, messages, max_tokens=4096)

    if err or not reply:
        for t in turns:
            t["topic"] = t["user_message"][:80] + ("..." if len(t["user_message"]) > 80 else "")
            t["summary"] = t["user_message"][:200]
            t["category"] = _guess_conversation_category(t["user_message"])
    else:
        import json as _json
        try:
            clean = reply.strip()
            if clean.startswith("```"):
                clean = clean.split("\n", 1)[1].rsplit("```", 1)[0].strip()
            summaries = _json.loads(clean)
            for s in summaries:
                idx = s.get("index", -1)
                if 0 <= idx < len(turns):
                    turns[idx]["topic"] = s.get("topic", turns[idx]["user_message"][:80])
                    turns[idx]["summary"] = s.get("summary", turns[idx]["user_message"][:200])
                    turns[idx]["category"] = s.get("category", "General")
        except Exception:
            for t in turns:
                t["topic"] = t["user_message"][:80] + ("..." if len(t["user_message"]) > 80 else "")
                t["summary"] = t["user_message"][:200]
                t["category"] = _guess_conversation_category(t["user_message"])

    for t in turns:
        t.setdefault("topic", t["user_message"][:80])
        t.setdefault("summary", t["user_message"][:200])
        t.setdefault("category", _guess_conversation_category(t["user_message"]))

    existing.extend(turns)
    save_pending_digest(session_id, existing)
    return {"ok": True, "generated": len(turns), "message": f"Generated {len(turns)} digest items."}


@app.post("/api/conversation/digest/review")
def api_digest_review(body: dict):
    """PCM approves or discards digest items.
    body: {session_id, approve: [ids], discard: [ids]}
    """
    session_id = body.get("session_id", "")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")

    approve_ids = body.get("approve", [])
    discard_ids = body.get("discard", [])

    from schedule_agent_web.store import approve_digest_items, discard_digest_items

    vectorized = []
    if approve_ids:
        approved = approve_digest_items(session_id, approve_ids)
        from schedule_agent_web.vector_store import index_conversation_turn, is_qdrant_available
        from schedule_agent_web.store import _load_json, _save_json, _approved_path
        if is_qdrant_available():
            for item in approved:
                success = index_conversation_turn(
                    session_id,
                    item.get("user_message", ""),
                    item.get("assistant_response", ""),
                )
                item["vectorized"] = bool(success)
                if success:
                    vectorized.append(item.get("id"))
        else:
            for item in approved:
                item["vectorized"] = False

        approved_log = _load_json(_approved_path(session_id))
        approved_ids_set = {i.get("id") for i in approved}
        for stored in approved_log:
            match = next((a for a in approved if a.get("id") == stored.get("id")), None)
            if match:
                stored["vectorized"] = match.get("vectorized", False)
        _save_json(_approved_path(session_id), approved_log)

    discarded_count = 0
    if discard_ids:
        discarded_count = discard_digest_items(session_id, discard_ids)

    return {
        "ok": True,
        "approved": len(approve_ids),
        "vectorized": len(vectorized),
        "discarded": discarded_count,
    }


@app.post("/api/conversation/vectorize")
def api_conversation_vectorize(body: dict):
    """Vectorize an already-approved conversation item."""
    session_id = body.get("session_id", "")
    item_id = body.get("item_id", "")
    if not session_id or not item_id:
        raise HTTPException(status_code=400, detail="session_id and item_id required")

    from schedule_agent_web.store import _load_json, _save_json, _approved_path
    from schedule_agent_web.vector_store import index_conversation_turn, is_qdrant_available

    if not is_qdrant_available():
        raise HTTPException(status_code=503, detail="Vector store (Qdrant) is not available. Start Qdrant and try again.")

    approved_log = _load_json(_approved_path(session_id))
    item = next((i for i in approved_log if i.get("id") == item_id), None)
    if not item:
        raise HTTPException(status_code=404, detail="Approved item not found.")

    success = index_conversation_turn(
        session_id,
        item.get("user_message", ""),
        item.get("assistant_response", ""),
    )
    item["vectorized"] = bool(success)
    _save_json(_approved_path(session_id), approved_log)

    if not success:
        raise HTTPException(status_code=500, detail="Vectorization failed.")
    return {"ok": True, "vectorized": True}


@app.get("/api/conversation/discarded")
def api_discarded_log(session_id: str = ""):
    """Get discarded conversation items (30-day retention)."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import get_discarded_log
    return {"ok": True, "items": get_discarded_log(session_id)}


@app.get("/api/conversation/approved")
def api_approved_log(session_id: str = ""):
    """Get approved (vectorized) conversation items."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import get_approved_log
    return {"ok": True, "items": get_approved_log(session_id)}


@app.get("/api/library")
def api_library_list(session_id: str = ""):
    """List all project library files with category and vectorization status."""
    if not session_id:
        return {"files": []}
    from schedule_agent_web.store import get_files
    all_files = get_files(session_id)
    library = []
    seen_display: set[str] = set()
    for f in all_files:
        fname = f.get("filename", "")
        if fname.startswith("_"):
            continue
        display = _library_display_name(fname) or fname
        if display in seen_display:
            continue
        seen_display.add(display)
        cat = f.get("category") or _library_auto_category(fname) or "Uncategorized"
        entry = {
            "filename": fname,
            "size": f.get("size", 0),
            "uploaded_at": f.get("uploaded_at", ""),
            "category": cat,
            "vectorized": f.get("vectorized", False),
        }
        if display != fname:
            entry["display_name"] = display
        library.append(entry)
    return {"files": library}


class LibraryUploadRequest(BaseModel):
    session_id: str
    filename: str
    content: str
    category: str = ""


@app.post("/api/library/upload")
def api_library_upload(request: LibraryUploadRequest):
    """Upload a file to the project library with a category. Vectorizes automatically."""
    if not request.session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    if not request.filename:
        raise HTTPException(status_code=400, detail="filename required")
    if not (request.category or "").strip():
        raise HTTPException(status_code=400, detail="Category is required.")
    category = request.category.strip()

    from schedule_agent_web.store import save_file, update_file_meta
    from schedule_agent_web.upload_utils import process_upload_content
    is_pdf = request.filename.lower().endswith(".pdf")
    if is_pdf:
        content_to_store = process_upload_content(request.filename, request.content)
    else:
        content_to_store = request.content

    result = save_file(request.session_id, request.filename, content_to_store)
    if not result:
        raise HTTPException(status_code=500, detail="Failed to save file")

    vectorized = False
    try:
        from schedule_agent_web.vector_store import index_file
        vectorized = index_file(request.session_id, request.filename, content_to_store)
    except Exception:
        pass

    update_file_meta(request.session_id, request.filename, category=category, vectorized=vectorized)
    result["category"] = category
    result["vectorized"] = vectorized
    return {"ok": True, "file": result}


def _is_user_visible_internal(filename: str) -> bool:
    """Check if a _-prefixed file is user-visible (legacy support)."""
    import re
    return bool(re.match(r"^_baseline_v\d+_", filename) or filename.startswith("_specs_"))


@app.put("/api/library/category")
def api_library_update_category(session_id: str = "", filename: str = "", category: str = ""):
    """Update the category of an existing library file."""
    if not session_id or not filename:
        raise HTTPException(status_code=400, detail="session_id and filename required")
    if filename.startswith("_") and not _is_user_visible_internal(filename):
        raise HTTPException(status_code=403, detail="Cannot modify internal files.")
    from schedule_agent_web.store import update_file_meta
    ok = update_file_meta(session_id, filename, category=(category or "").strip() or "Uncategorized")
    if not ok:
        raise HTTPException(status_code=404, detail="File not found")
    return {"ok": True}


@app.delete("/api/library")
def api_library_delete(session_id: str = "", filename: str = ""):
    """Delete a file from the project library and remove its vectors."""
    if not session_id or not filename:
        raise HTTPException(status_code=400, detail="session_id and filename required")
    if filename.startswith("_") and not _is_user_visible_internal(filename):
        raise HTTPException(status_code=403, detail="Cannot delete internal files.")
    from schedule_agent_web.store import delete_file
    if not delete_file(session_id, filename):
        raise HTTPException(status_code=404, detail="File not found")
    try:
        from schedule_agent_web.vector_store import delete_file_vectors
        delete_file_vectors(session_id, filename)
    except Exception:
        pass
    try:
        from schedule_agent_web.baseline import delete_submissions_by_file
        delete_submissions_by_file(session_id, filename)
    except Exception:
        pass
    return {"ok": True, "message": f"'{filename}' removed from project library and vector store."}


class RevectorizeRequest(BaseModel):
    session_id: str
    filename: str


@app.post("/api/library/revectorize")
def api_library_revectorize(request: RevectorizeRequest):
    """Re-vectorize an existing library file (e.g., after vector store reset)."""
    session_id = request.session_id
    filename = request.filename
    if not session_id or not filename:
        raise HTTPException(status_code=400, detail="session_id and filename required")
    if filename.startswith("_") and not _is_user_visible_internal(filename):
        raise HTTPException(status_code=403, detail="Cannot modify internal files.")
    from schedule_agent_web.store import get_file_content, update_file_meta
    content = get_file_content(session_id, filename)
    if not content:
        raise HTTPException(status_code=404, detail="File content not found")
    vectorized = False
    vec_error = ""
    try:
        from schedule_agent_web.vector_store import delete_file_vectors, index_file, is_qdrant_available
        if not is_qdrant_available():
            vec_error = "Vector store (Qdrant) is not configured or unreachable."
        else:
            delete_file_vectors(session_id, filename)
            vectorized = index_file(session_id, filename, content)
            if not vectorized:
                vec_error = "Embedding failed. Check OPENAI_API_KEY or sentence-transformers installation."
    except Exception as e:
        vec_error = str(e)
    update_file_meta(session_id, filename, vectorized=vectorized)
    result = {"ok": True, "vectorized": vectorized}
    if vec_error:
        result["warning"] = vec_error
    return result


@app.get("/api/library/verify")
def api_library_verify(session_id: str = ""):
    """Verify every file's vectorized status against Qdrant and correct metadata if needed."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.store import get_files, update_file_meta
    from schedule_agent_web.vector_store import count_file_vectors, is_qdrant_available

    qdrant_up = is_qdrant_available()

    all_files = get_files(session_id)
    results = []
    corrected = 0
    for f in all_files:
        fname = f.get("filename", "")
        meta_says = f.get("vectorized", False)
        if qdrant_up:
            count = count_file_vectors(session_id, fname)
        else:
            count = -1
        actually_vectorized = count > 0 if count >= 0 else False

        entry = {
            "filename": fname,
            "meta_vectorized": meta_says,
            "qdrant_vectors": count,
            "actually_vectorized": actually_vectorized,
            "status": "ok",
        }

        if not qdrant_up:
            if meta_says:
                update_file_meta(session_id, fname, vectorized=False)
                entry["status"] = "corrected_to_false"
                corrected += 1
            else:
                entry["status"] = "qdrant_unavailable"
        elif meta_says and not actually_vectorized:
            update_file_meta(session_id, fname, vectorized=False)
            entry["status"] = "corrected_to_false"
            corrected += 1
        elif not meta_says and actually_vectorized:
            update_file_meta(session_id, fname, vectorized=True)
            entry["status"] = "corrected_to_true"
            corrected += 1

        results.append(entry)

    return {
        "ok": True,
        "qdrant_available": qdrant_up,
        "total_files": len(results),
        "corrected": corrected,
        "files": results,
    }


# --- Admin-only: internal/system file management ---

def _require_admin(authorization: str) -> dict:
    """Verify token and require admin role. Raises HTTPException if not admin."""
    from schedule_agent_web.auth import verify_token
    token = authorization.replace("Bearer ", "") if authorization.startswith("Bearer ") else authorization
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    user = verify_token(token)
    if user is None:
        raise HTTPException(status_code=401, detail="Invalid or expired token")
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


@app.get("/api/admin/files")
def api_admin_list_files(session_id: str = "", authorization: str = Header(default="")):
    """Admin only: list ALL files including internal ones."""
    _require_admin(authorization)
    if not session_id:
        return {"files": []}
    from schedule_agent_web.store import get_files
    all_files = get_files(session_id)
    return {
        "files": [
            {
                "filename": f.get("filename", ""),
                "size": f.get("size", 0),
                "uploaded_at": f.get("uploaded_at", ""),
                "category": f.get("category", ""),
                "vectorized": f.get("vectorized", False),
                "internal": f.get("filename", "").startswith("_"),
            }
            for f in all_files
        ]
    }


@app.delete("/api/admin/files")
def api_admin_delete_file(session_id: str = "", filename: str = "", authorization: str = Header(default="")):
    """Admin only: delete ANY file including internal ones."""
    _require_admin(authorization)
    if not session_id or not filename:
        raise HTTPException(status_code=400, detail="session_id and filename required")
    from schedule_agent_web.store import delete_file
    if not delete_file(session_id, filename):
        raise HTTPException(status_code=404, detail="File not found")
    try:
        from schedule_agent_web.vector_store import delete_file_vectors
        delete_file_vectors(session_id, filename)
    except Exception:
        pass
    return {"ok": True, "message": f"'{filename}' deleted by admin."}


class FileUploadRequest(BaseModel):
    session_id: str
    filename: str
    content: str = ""
    category: str = ""


@app.post("/api/upload")
async def api_upload_file(request: FileUploadRequest):
    """Upload a file (CSV, MD, TXT, XER, PDF scope of work, etc.) for this session. Max 10MB. Text or base64 for PDF."""
    if not request.session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    if not request.filename:
        raise HTTPException(status_code=400, detail="filename required")
    try:
        from schedule_agent_web.upload_utils import process_upload_content
        is_pdf = request.filename.lower().endswith(".pdf")
        if is_pdf:
            raw = (request.content or "").strip()
            if raw.startswith("data:"):
                raw = raw.split(",", 1)[-1]
            import base64
            try:
                pdf_bytes = base64.b64decode(raw, validate=True)
            except Exception:
                raise HTTPException(status_code=400, detail="Invalid base64 for PDF")
            if len(pdf_bytes) > 500 * 1024 * 1024:
                raise HTTPException(status_code=400, detail="File too large (max 500MB)")
            content_to_store = process_upload_content(request.filename, request.content)
        else:
            content_bytes = request.content.encode("utf-8") if isinstance(request.content, str) else request.content
            if len(content_bytes) > 500 * 1024 * 1024:
                raise HTTPException(status_code=400, detail="File too large (max 500MB)")
            content_to_store = request.content

        from schedule_agent_web.store import save_file
        result = save_file(request.session_id, request.filename, content_to_store)
        if not result:
            raise HTTPException(status_code=500, detail="Failed to save file")
        vectorized = False
        try:
            from schedule_agent_web.vector_store import index_file
            vectorized = index_file(request.session_id, request.filename, content_to_store)
        except Exception:
            pass
        from schedule_agent_web.store import update_file_meta
        cat = (request.category or "").strip() or _auto_category_from_filename(request.filename)
        update_file_meta(request.session_id, request.filename, vectorized=vectorized, category=cat)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _auto_category_from_filename(filename: str) -> str:
    """Best-guess category from filename and extension."""
    lower = filename.lower()
    ext = lower.rsplit(".", 1)[-1] if "." in lower else ""
    if ext == "xer":
        return "Baseline Schedule (XER)"
    if any(kw in lower for kw in ["spec", "provision", "technical"]):
        return "Contract Specifications"
    if any(kw in lower for kw in ["narrative", "baseline schedule"]):
        return "Baseline Narrative"
    if any(kw in lower for kw in ["scope", "sow"]):
        return "Scope Documents"
    if any(kw in lower for kw in ["rfi", "submittal"]):
        return "RFI / Submittals"
    if any(kw in lower for kw in ["change order", "co_", "changeorder"]):
        return "Change Orders"
    if any(kw in lower for kw in ["meeting minute", "mom_", "minutes"]):
        return "Meeting Minutes"
    if any(kw in lower for kw in ["daily", "field log", "daily report"]):
        return "Daily Field Logs"
    if any(kw in lower for kw in ["cost", "budget", "estimate"]):
        return "Cost Reports"
    if any(kw in lower for kw in ["risk", "qsra", "qcra"]):
        return "Risk Register"
    if any(kw in lower for kw in ["safety", "loto", "jha"]):
        return "Safety Documents"
    if any(kw in lower for kw in ["commission", "turnover", "punchlist"]):
        return "Commissioning"
    if any(kw in lower for kw in ["contract", "agreement"]):
        return "Contract Documents"
    return "Uncategorized"


# ── Contractor Baseline Review API ────────────────────────

class BaselineSubmitRequest(BaseModel):
    session_id: str
    submission_type: str = "baseline"
    version: int | None = None
    xer_filename: str | None = None
    xer_content: str | None = None
    xer_size: int = 0
    narr_filename: str | None = None
    narr_content: str | None = None
    narr_size: int = 0
    resp_filename: str | None = None
    resp_content: str | None = None
    resp_size: int = 0


class BaselineSpecsRequest(BaseModel):
    session_id: str
    filename: str
    content: str
    size: int = 0


@app.get("/api/baseline/status")
def api_baseline_status(session_id: str = ""):
    """Return specs status, scope-doc availability, and submission history."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.baseline import get_specs, list_submissions, has_scope_docs, save_specs_meta
    from schedule_agent_web.store import get_files

    specs = get_specs(session_id)
    scope_docs_available = has_scope_docs(session_id)
    library_specs = []

    store_files = get_files(session_id) if scope_docs_available else []
    _scope_cats = {"Contract Specifications", "Scope Documents", "Contract Documents"}
    _baseline_cats = {"Baseline Schedule (XER)", "Baseline Narrative", "Baseline Submission"}
    _update_cats = {"Schedule Update", "Schedule Update (XER)", "Update Narrative"}

    def _file_entry(f):
        return {"filename": f.get("filename", ""), "size": f.get("size", 0),
                "uploaded_at": f.get("uploaded_at", ""), "category": f.get("category", "")}

    all_project_files = [f for f in store_files if not f.get("filename", "").startswith("_")]
    scope_file_list = [_file_entry(f) for f in all_project_files if f.get("category", "") in _scope_cats]
    baseline_file_list = [_file_entry(f) for f in all_project_files if f.get("category", "") in _baseline_cats]
    update_file_list = [_file_entry(f) for f in all_project_files if f.get("category", "") in _update_cats]

    if not specs:
        try:
            from schedule_agent_web.ingestion import list_ingested_documents, get_ingested_document
            ingested = list_ingested_documents(session_id, ingestion_group="contract_specs")
            if ingested:
                for doc in ingested:
                    fname = doc.get("filename", "")
                    library_specs.append({
                        "filename": fname,
                        "doc_id": doc.get("id", ""),
                        "size": (doc.get("metadata") or {}).get("size_bytes", 0),
                        "uploaded_at": (doc.get("metadata") or {}).get("ingested_at", ""),
                    })
                spec_doc = ingested[0]
                spec_fname = spec_doc.get("filename", "")
                spec_size = (spec_doc.get("metadata") or {}).get("size_bytes", 0)
                full_doc = get_ingested_document(session_id, spec_doc.get("id", ""))
                if full_doc and full_doc.raw_text:
                    from schedule_agent_web.store import save_file
                    save_file(session_id, spec_fname, full_doc.raw_text)
                    save_specs_meta(session_id, spec_fname, spec_size)
                    specs = get_specs(session_id)
                    if specs:
                        specs["from_library"] = True
        except Exception:
            pass

    if not specs and not library_specs and scope_file_list:
        spec_candidates = [
            f for f in scope_file_list
            if any(kw in f["filename"].lower() for kw in
                   ["spec", "provision", "technical", "scope", "contract"])
        ]
        if not spec_candidates:
            spec_candidates = scope_file_list
        if spec_candidates:
            chosen = spec_candidates[0]
            from schedule_agent_web.store import get_file_content
            content = get_file_content(session_id, chosen["filename"])
            if content:
                from schedule_agent_web.store import save_file
                save_file(session_id, chosen["filename"], content)
                save_specs_meta(session_id, chosen["filename"], chosen.get("size", 0))
                specs = get_specs(session_id)
                if specs:
                    specs["from_library"] = True
            library_specs = spec_candidates

    from schedule_agent_web.baseline import next_expected_version
    baseline_subs = list_submissions(session_id, "baseline")
    update_subs = list_submissions(session_id, "update")
    baseline_next_ver = next_expected_version(session_id, "baseline")
    update_next_ver = next_expected_version(session_id, "update")

    baseline_complete = False
    if baseline_subs:
        latest_bl = baseline_subs[0]
        baseline_complete = bool(latest_bl.get("xer_filename")) and bool(latest_bl.get("narr_filename"))

    return {
        "specs": specs,
        "has_scope_docs": scope_docs_available,
        "library_specs": library_specs,
        "scope_file_list": scope_file_list,
        "baseline_file_list": baseline_file_list,
        "update_file_list": update_file_list,
        "baseline_submissions": baseline_subs,
        "update_submissions": update_subs,
        "baseline_next_version": baseline_next_ver,
        "update_next_version": update_next_ver,
        "baseline_complete": baseline_complete,
    }


@app.post("/api/baseline/specs")
def api_baseline_specs(request: BaselineSpecsRequest):
    """Upload contract specifications (one-time, immutable)."""
    if not request.session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    if not request.filename:
        raise HTTPException(status_code=400, detail="filename required")
    from schedule_agent_web.baseline import get_specs, save_specs_meta
    existing = get_specs(request.session_id)
    if existing:
        raise HTTPException(status_code=409, detail="Contract specs already uploaded. They cannot be replaced.")
    from schedule_agent_web.store import save_file
    save_file(request.session_id, request.filename, request.content)
    meta = save_specs_meta(request.session_id, request.filename, request.size)
    try:
        from schedule_agent_web.vector_store import index_file
        index_file(request.session_id, request.filename, request.content)
    except Exception:
        pass
    return {"ok": True, "specs": meta}


@app.post("/api/baseline/submit")
def api_baseline_submit(request: BaselineSubmitRequest):
    """Submit a new baseline/update version or attach files to an existing version."""
    if not request.session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    stype = request.submission_type if request.submission_type in ("baseline", "update") else "baseline"
    prefix = "baseline" if stype == "baseline" else "update"
    xer_cat = "Baseline Schedule (XER)" if stype == "baseline" else "Schedule Update (XER)"
    narr_cat = "Baseline Narrative" if stype == "baseline" else "Update Narrative"

    from schedule_agent_web.baseline import (
        get_specs, has_scope_docs, create_submission, update_submission,
        list_submissions, get_submission,
    )
    from schedule_agent_web.store import save_file, update_file_meta

    specs = get_specs(request.session_id)
    scope = has_scope_docs(request.session_id)
    if not specs and not scope:
        raise HTTPException(
            status_code=422,
            detail="Upload contract specs or scope documents before submitting.",
        )

    if not request.xer_filename and not request.narr_filename and not request.resp_filename:
        raise HTTPException(status_code=400, detail="At least one file (XER, narrative, or response) is required.")

    def _save_files(ver):
        if request.xer_filename and request.xer_content:
            if not request.xer_filename.lower().endswith(".xer"):
                raise HTTPException(status_code=400, detail="XER file must have .xer extension")
            xer_store = f"{prefix}_v{ver}_{request.xer_filename}"
            save_file(request.session_id, xer_store, request.xer_content)
            update_file_meta(request.session_id, xer_store, category=xer_cat)
        if request.narr_filename and request.narr_content:
            narr_store = f"{prefix}_v{ver}_{request.narr_filename}"
            save_file(request.session_id, narr_store, request.narr_content)
            update_file_meta(request.session_id, narr_store, category=narr_cat)
        if request.resp_filename and request.resp_content:
            resp_store = f"{prefix}_v{ver}_resp_{request.resp_filename}"
            save_file(request.session_id, resp_store, request.resp_content)
            update_file_meta(request.session_id, resp_store, category="Contractor Response")

    if request.version:
        existing = get_submission(request.session_id, request.version, stype)
        if not existing:
            raise HTTPException(status_code=404, detail=f"{stype.title()} version {request.version} not found.")
        ver = request.version
        updates = {}
        if request.xer_filename and request.xer_content:
            updates["xer_filename"] = request.xer_filename
            updates["xer_size"] = request.xer_size
        if request.narr_filename and request.narr_content:
            updates["narr_filename"] = request.narr_filename
            updates["narr_size"] = request.narr_size
        if request.resp_filename and request.resp_content:
            updates["resp_filename"] = request.resp_filename
            updates["resp_size"] = request.resp_size
        _save_files(ver)
        sub = update_submission(request.session_id, ver, stype, **updates)
        return {"ok": True, "submission": sub, "updated": True}

    if not request.xer_filename and not request.narr_filename:
        raise HTTPException(status_code=400, detail="XER or narrative file required for a new submission.")
    if request.xer_filename and not request.xer_filename.lower().endswith(".xer"):
        raise HTTPException(status_code=400, detail="XER file must have .xer extension")
    existing_subs = list_submissions(request.session_id, stype)
    is_subsequent = len(existing_subs) > 0
    if is_subsequent and not request.resp_filename:
        raise HTTPException(
            status_code=422,
            detail="Contractor's filled comment response file is mandatory for v2+ submissions.",
        )
    sub = create_submission(
        session_id=request.session_id,
        submission_type=stype,
        xer_filename=request.xer_filename or "",
        xer_size=request.xer_size,
        narr_filename=request.narr_filename,
        narr_size=request.narr_size,
        resp_filename=request.resp_filename,
        resp_size=request.resp_size,
    )
    ver = sub["version"]
    _save_files(ver)
    return {"ok": True, "submission": sub}


def _decode_xer_content(raw: str) -> str:
    """If content looks like base64 (from legacy binary upload), decode it to text."""
    import base64, re
    if not raw or raw.startswith("%T") or raw.startswith("ERMHDR"):
        return raw
    if re.match(r'^[A-Za-z0-9+/\r\n]+=*$', raw[:200].replace('\n', '').replace('\r', '')):
        try:
            decoded = base64.b64decode(raw).decode("utf-8", errors="replace")
            if "%T" in decoded[:200] or "ERMHDR" in decoded[:200]:
                return decoded
        except Exception:
            pass
    return raw


def _parse_xer_tables(xer_text: str) -> dict[str, list[dict]]:
    """Parse XER text into named tables. Returns {table_name: [row_dicts]}."""
    tables: dict[str, list[dict]] = {}
    current_table = None
    headers: list[str] = []

    for line in xer_text.split("\n"):
        line = line.rstrip("\r")
        if line.startswith("%T\t"):
            current_table = line.split("\t", 1)[1].strip()
            headers = []
            tables[current_table] = []
        elif line.startswith("%F\t") and current_table:
            headers = [h.strip() for h in line.split("\t")[1:]]
        elif line.startswith("%R\t") and current_table and headers:
            vals = line.split("\t")[1:]
            row = {}
            for i, h in enumerate(headers):
                row[h] = vals[i].strip() if i < len(vals) else ""
            tables[current_table].append(row)
        elif line.startswith("%E"):
            current_table = None
            headers = []

    return tables


def _summarize_xer_for_review(xer_text: str) -> str:
    """Parse XER and produce a structured summary focused on activities, logic, and WBS."""
    tables = _parse_xer_tables(xer_text)
    parts = []

    wbs = tables.get("PROJWBS", [])
    if wbs:
        parts.append(f"## WBS Structure ({len(wbs)} entries)")
        for w in wbs[:200]:
            wbs_id = w.get("wbs_id", "")
            short = w.get("wbs_short_name", "")
            name = w.get("wbs_name", "")
            parent = w.get("parent_wbs_id", "")
            parts.append(f"  WBS_ID={wbs_id}  SHORT={short}  NAME={name}  PARENT={parent}")

    tasks = tables.get("TASK", [])
    if tasks:
        parts.append(f"\n## Activities ({len(tasks)} tasks)")
        cols = ["task_id", "task_code", "task_name", "wbs_id", "task_type",
                "total_float_hr_cnt", "duration_type", "target_drtn_hr_cnt",
                "remain_drtn_hr_cnt", "target_start_date", "target_end_date",
                "status_code", "phys_complete_pct"]
        header_line = "\t".join(c for c in cols if any(c in t for t in tasks[:1]))
        actual_cols = [c for c in cols if tasks and c in tasks[0]]
        if actual_cols:
            parts.append("  " + "\t".join(actual_cols))
            for t in tasks:
                parts.append("  " + "\t".join(str(t.get(c, "")) for c in actual_cols))

    preds = tables.get("TASKPRED", [])
    if preds:
        parts.append(f"\n## Logic Relationships ({len(preds)} ties)")
        pred_cols = ["task_id", "pred_task_id", "pred_type", "lag_hr_cnt", "pred_type"]
        actual = [c for c in pred_cols if preds and c in preds[0]]
        if not actual:
            actual = list(preds[0].keys()) if preds else []
        parts.append("  " + "\t".join(actual))
        for p in preds:
            parts.append("  " + "\t".join(str(p.get(c, "")) for c in actual))

    rsrc = tables.get("TASKRSRC", [])
    if rsrc:
        parts.append(f"\n## Resource Assignments ({len(rsrc)} entries)")
        rsrc_cols = ["task_id", "rsrc_id", "target_qty", "target_cost", "remain_qty", "remain_cost"]
        actual = [c for c in rsrc_cols if rsrc and c in rsrc[0]]
        if actual:
            parts.append("  " + "\t".join(actual))
            for r in rsrc[:300]:
                parts.append("  " + "\t".join(str(r.get(c, "")) for c in actual))

    cal = tables.get("CALENDAR", [])
    if cal:
        parts.append(f"\n## Calendars ({len(cal)} entries)")
        for c in cal[:30]:
            parts.append(f"  CLNDR_ID={c.get('clndr_id','')}  NAME={c.get('clndr_name','')}")

    if not parts:
        return xer_text

    summary = "\n".join(parts)
    return summary


class ReviewExecuteRequest(BaseModel):
    session_id: str
    version: int
    columns: list[str] = []
    submission_type: str = "baseline"


@app.post("/api/baseline/review/execute")
def api_baseline_review_execute(request: ReviewExecuteRequest):
    """Execute AI-powered baseline review. Reads specs + XER + narrative, generates review comments."""
    if not request.session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    if not get_status_dict().get("has_api_key"):
        raise HTTPException(status_code=503, detail="Set OPENAI_API_KEY or ANTHROPIC_API_KEY in environment.")

    from schedule_agent_web.baseline import (
        get_submission, get_specs, get_review, save_review_result,
        save_review_meta, get_default_columns, get_last_comment_number,
    )
    from schedule_agent_web.store import get_file_content

    stype = request.submission_type if request.submission_type in ("baseline", "update") else "baseline"
    sub = get_submission(request.session_id, request.version, submission_type=stype)
    if not sub:
        raise HTTPException(status_code=404, detail=f"{stype.title()} v{request.version} not found.")
    existing = get_review(request.session_id, request.version)
    if existing:
        raise HTTPException(status_code=409, detail=f"Review already exists for v{request.version}. Download from results.")

    ver = request.version
    columns = request.columns if request.columns else get_default_columns(ver)
    file_prefix = stype  # "baseline" or "update"

    specs_content = ""
    specs_info = get_specs(request.session_id)
    if specs_info:
        specs_content = get_file_content(request.session_id, specs_info["filename"]) or ""

    xer_raw = get_file_content(request.session_id, f"{file_prefix}_v{ver}_{sub['xer_filename']}") or ""
    xer_content = _decode_xer_content(xer_raw)
    narr_content = ""
    if sub.get("narr_filename"):
        narr_content = get_file_content(request.session_id, f"{file_prefix}_v{ver}_{sub['narr_filename']}") or ""

    resp_content = ""
    if sub.get("resp_filename"):
        resp_content = get_file_content(request.session_id, f"{file_prefix}_v{ver}_resp_{sub['resp_filename']}") or ""

    has_prev_comments = ver > 1 and resp_content

    prev_last_num = get_last_comment_number(request.session_id)
    start_num = prev_last_num + 1

    col_list = ", ".join(columns)

    addressed_instruction = ""
    if has_prev_comments:
        addressed_instruction = (
            "- 'Addressed (Yes/No)' column: For EVERY new comment, you MUST also verify the contractor's "
            "responses from the previous version's filled comment sheet (provided below). "
            "For each new comment, if the contractor's response to a prior related comment adequately addresses "
            "the issue per the contract specifications, set 'Addressed (Yes/No)' to 'Yes'. "
            "If the response is insufficient, incomplete, or non-compliant with specs, set it to 'No'. "
            "If the comment is entirely new (not related to any prior comment), set it to null.\n"
        )
    else:
        addressed_instruction = (
            "- 'Addressed (Yes/No)' must be null for all comments (this is the first version or no contractor response available).\n"
        )

    review_label = "Baseline" if stype == "baseline" else "Update"
    system_prompt = (
        f"You are a senior CPM scheduling expert performing a Contractor {review_label} Schedule Review. "
        f"You are given contract specifications, "
        f"an XER schedule export ({review_label} v{ver}), and optionally a schedule narrative ({review_label} v{ver}). "
        f"The trio of XER + Narrative + Review Comments belongs to {review_label} v{ver}. "
        "Produce a thorough review with actionable comments.\n\n"
        f"Output ONLY a valid JSON array of objects. Each object must have these keys: {col_list}.\n"
        "Rules:\n"
        f"- 'Comment ID' must be sequential starting from BLR-{start_num:03d}. "
        f"Previous reviews used IDs up to BLR-{prev_last_num:03d}. "
        "Reuse of any prior Comment ID is strictly prohibited.\n"
        "- 'WBS Reference' should identify the WBS element or activity ID the comment relates to, or 'General' if project-wide.\n"
        "- 'Comment Description' should be specific and reference actual activities, logic, or spec requirements.\n"
        "- 'Spec Section Reference' should cite the specific contract spec section or 'General' if broadly applicable.\n"
        "- 'Priority' must be one of: 'Mandatory' or 'Recommendation'. 'Mandatory' for spec violations, missing scope, "
        "logic errors, DCMA failures. 'Recommendation' for best-practice improvements.\n"
        "\n## Logic Flag (CRITICAL — Logic Relationship Analysis)\n"
        "- 'Logic Flag' column classifies the logic finding. Must be one of:\n"
        "  - 'Missing Hard Logic' — A mandatory FS/SS/FF/SF relationship that MUST exist based on physical "
        "construction sequence or spec requirements is absent. Examples: concrete cure before formwork strip, "
        "backfill before paving, inspection before cover-up, equipment set before piping tie-in, "
        "testing before energization. These are physical dependencies that cannot be overridden.\n"
        "  - 'Preferential Cross-WBS' — A preferential (soft/resource-driven) logic tie that crosses WBS levels. "
        "These are NOT physically required but were added for resource leveling or sequencing preference. "
        "Cross-WBS preferential logic constrains float artificially and must be flagged for PCM review. "
        "Example: a finish-to-start tie from 'Civil Works' WBS to 'MEP Fit-Out' WBS that is resource-based, "
        "not physically required.\n"
        "  - 'Dangling Logic' — An activity has no predecessor, no successor, or both (open-ended). "
        "Per DCMA 14-Point, activities should not be dangling except project start/finish milestones.\n"
        "  - 'Redundant Logic' — A relationship that is already implied by transitive logic through "
        "intermediate activities and adds no scheduling value.\n"
        "  - 'Incorrect Relationship Type' — The relationship type (FS/SS/FF/SF) does not match the physical "
        "or contractual requirement. Example: SS used where FS is physically required.\n"
        "  - '' (empty string) — Comment is not logic-related.\n"
        "\nLogic review requirements:\n"
        "- Analyze the XER TASKPRED table (or relationship data) for all predecessor/successor ties.\n"
        "- For EVERY activity, check if required hard logic connections exist based on construction sequence "
        "and spec requirements. Flag missing ones as 'Missing Hard Logic' with Priority 'Mandatory'.\n"
        "- For EVERY relationship that crosses WBS levels (predecessor and successor belong to different WBS), "
        "determine if it is physically required (hard) or preferential (soft/resource). "
        "Flag preferential cross-WBS ties as 'Preferential Cross-WBS' with Priority 'Recommendation'.\n"
        "- Check for dangling activities (no predecessor or no successor).\n"
        "- Ensure at least 5-10 logic-specific comments are included in the review.\n\n"
        "- 'Recommendation for Correction' should be clear, actionable guidance.\n"
        "- 'Contractor Response' must always be empty string (for contractor to fill in).\n"
        "- 'Comment Status' must be 'Open' for all comments.\n"
        + addressed_instruction +
        "- For any additional custom columns, provide appropriate values.\n"
        "- Review against DCMA 14-Point Assessment, CPM best practices, spec compliance, logic integrity, "
        "duration reasonableness, float analysis, missing scope, resource loading, and cost loading.\n"
        "- There is NO LIMIT on the number of comments. Generate as many comments as the schedule warrants. "
        "Every spec violation, logic deficiency, missing scope item, DCMA failure, duration issue, float concern, "
        "and best-practice deviation should be its own comment. A thorough review of a real schedule typically "
        "produces 30-100+ comments. Do NOT truncate or summarize — every finding must be a separate comment.\n"
        "- Output ONLY the JSON array, no markdown, no explanation."
    )

    # Inject project-specific philosophy settings into the review prompt
    try:
        from schedule_agent_web.store import get_philosophy
        phil = get_philosophy(request.session_id)
        gov = phil.get("governance", {})
        bas = phil.get("basis", {})
        pb = phil.get("playbook", {})
        phil_instructions = "\n\n## Project-Specific Review Configuration (MUST APPLY)\n"
        std_map = {"dcma": "DCMA 14-Point", "aace": "AACE International", "custom": "Owner Custom"}
        tone_map = {"exec": "Executive — formal, concise, suitable for owner leadership", "tech": "Technical — detailed engineering language", "aggr": "Aggressive — flag every issue firmly, no soft language", "cons": "Conservative — balanced, diplomatic tone"}
        phil_instructions += f"Compliance standard: {std_map.get(gov.get('complianceStandard', 'dcma'), 'DCMA 14-Point')}\n"
        phil_instructions += f"Missing logic tolerance: {gov.get('missingLogicTolerance', 5)}% — only flag if missing logic exceeds this.\n"
        phil_instructions += f"High float threshold: {gov.get('highFloatDays', 44)} days — flag activities with total float above this.\n"
        phil_instructions += f"High duration threshold: {gov.get('highDurationDays', 20)} days — flag activities with duration above this as potentially needing decomposition.\n"
        phil_instructions += f"Negative lag tolerance: {gov.get('negativeLagTolerance', 0)} — maximum number of negative lags (leads) allowed before flagging.\n"
        phil_instructions += f"Hard constraint tolerance: {gov.get('hardConstraintTolerance', 5)}% — flag if hard constraints exceed this percentage of activities.\n"
        rules = gov.get("rules", {})
        if rules.get("leadRestriction"): phil_instructions += "RULE: Negative lags (leads) are FORBIDDEN. Flag every lead as Mandatory deficiency.\n"
        if rules.get("sfBan"): phil_instructions += "RULE: Start-to-Finish (SF) relationships are BANNED. Flag every SF as Mandatory deficiency.\n"
        if rules.get("hardConstraintAudit"): phil_instructions += "RULE: Hard constraints are discouraged. Every Must-Start-On, Must-Finish-On constraint must be flagged.\n"
        if rules.get("calendarCheck"): phil_instructions += "RULE: Every activity MUST have an assigned calendar. Missing calendars are Mandatory deficiencies.\n"
        if rules.get("fsPreferred"): phil_instructions += "RULE: FS relationships should be ≥ 80% of all relationships. Flag if percentage is lower.\n"
        phil_instructions += f"\nNarrative tone for all comments: {tone_map.get(gov.get('narrativeTone', 'exec'), 'Executive')}\n"
        phil_instructions += f"\nDelivery method: {bas.get('deliveryMethod', 'Design-Build')} — tailor review expectations accordingly.\n"
        if bas.get("ntpDate"): phil_instructions += f"NTP date: {bas['ntpDate']}\n"
        if bas.get("completionDate"): phil_instructions += f"Contractual completion: {bas['completionDate']}\n"
        phil_instructions += f"WBS structure: {bas.get('wbsDriver', 'area')}-driven, {bas.get('wbsLevels', 4)} levels\n"
        phil_instructions += f"Calendar assumptions: {bas.get('workWeek', '5-day')} week, {bas.get('hoursPerDay', 8)} hrs/day, {bas.get('shiftsPerDay', 1)} shift(s)\n"
        std_rev = {"approve-noted": "Approve as Noted — prefer approving with comments, list deficiencies that must be corrected", "approve-full": "Full Approval Required — do not recommend approval unless all deficiencies are zero", "reject-resubmit": "Reject & Resubmit — recommend rejection when deficiency count is high"}
        phil_instructions += f"\nReview standard: {std_rev.get(pb.get('reviewStandard', 'approve-noted'), 'Approve as Noted')}\n"
        phil_instructions += f"Rejection threshold: reject if deficiencies exceed {pb.get('maxDefectsBeforeReject', 20)} or missing scope > {pb.get('missingScopeThreshold', 10)}%\n"
        cl = pb.get("checklist", {})
        required_checks = [k.replace("_", " ").title() for k, v in cl.items() if v]
        if required_checks:
            phil_instructions += "Required review areas (must be covered): " + ", ".join(required_checks) + "\n"
        oa = pb.get("ownerActivities", [])
        req_oa = [a for a in oa if a.get("required")]
        if req_oa:
            phil_instructions += "Owner activities that MUST appear in the schedule: " + ", ".join(f"{a['name']} ({a['duration']}d)" for a in req_oa) + ". Flag as Mandatory deficiency if missing.\n"
        phil_instructions += "\nAt the END of the review, add a summary comment (last item) with 'WBS Reference': 'Summary' that provides:\n"
        phil_instructions += "- Overall recommendation (Approve as Noted / Reject for Resubmission) based on the review standard and thresholds above.\n"
        phil_instructions += "- Count of Mandatory vs Recommendation findings.\n"
        phil_instructions += "- Key areas of concern.\n"
        system_prompt += phil_instructions
    except Exception:
        pass

    user_msg_parts = []
    if specs_content:
        preview = specs_content[:12000]
        user_msg_parts.append(f"## Contract Specifications\n{preview}")
    if resp_content:
        preview = resp_content[:10000]
        user_msg_parts.append(
            f"## Contractor's Filled Comment Response (from previous version)\n"
            f"Verify each contractor response for compliance with contract specifications. "
            f"Mark 'Addressed (Yes/No)' accordingly.\n{preview}"
        )
    if xer_content:
        xer_summary = _summarize_xer_for_review(xer_content)
        xer_preview = xer_summary[:80000] if len(xer_summary) > 80000 else xer_summary
        user_msg_parts.append(f"## XER Schedule ({review_label} v{ver})\n{xer_preview}")
    if narr_content:
        preview = narr_content[:8000]
        user_msg_parts.append(f"## Schedule Narrative ({review_label} v{ver})\n{preview}")
    if not user_msg_parts:
        raise HTTPException(status_code=422, detail="No file content available for review.")

    user_msg = "\n\n".join(user_msg_parts)
    user_msg += f"\n\nGenerate the baseline review comments as a JSON array with columns: {col_list}"

    reply, err = _call_llm(system_prompt, [{"role": "user", "content": user_msg}], max_tokens=65536)
    if err:
        raise HTTPException(status_code=502, detail=f"AI review failed: {err}")

    comments = _extract_json_array(reply)
    if comments is None:
        retry_msg = (
            "Your previous response could not be parsed as valid JSON. "
            "Please output ONLY a valid JSON array (starting with [ and ending with ]). "
            "No markdown, no backticks, no explanation — just the JSON array."
        )
        reply2, err2 = _call_llm(system_prompt, [
            {"role": "user", "content": user_msg},
            {"role": "assistant", "content": reply},
            {"role": "user", "content": retry_msg},
        ], max_tokens=65536)
        if not err2:
            comments = _extract_json_array(reply2)
    if comments is None:
        raise HTTPException(status_code=502, detail="AI returned invalid JSON. Try again.")

    for idx, c in enumerate(comments):
        correct_num = start_num + idx
        c["Comment ID"] = f"BLR-{correct_num:03d}"
    last_comment_num = start_num + len(comments) - 1 if comments else prev_last_num

    source_files = {
        "xer": sub.get("xer_filename", ""),
        "specs": specs_info["filename"] if specs_info else "",
        "narrative": sub.get("narr_filename", ""),
        "response": sub.get("resp_filename", ""),
    }
    filepath = save_review_result(request.session_id, ver, comments, columns, source_files=source_files)

    exception_count = 0
    exception_filepath = ""
    exceptions = []
    if has_prev_comments:
        from schedule_agent_web.baseline import save_exception_report
        exc_system = (
            "You are a senior CPM scheduling expert performing a PCM Exception Review. "
            "Your task: identify DISCREPANCIES where the contractor claims they addressed a prior review comment "
            "(marked 'Yes' or claimed as addressed in their response) BUT the evidence in the new P6 XER file "
            "shows the issue was NOT actually corrected.\n\n"
            "You are given:\n"
            "1. Contract Specifications (the compliance standard)\n"
            "2. The contractor's filled comment response sheet (from the previous review version)\n"
            "3. The NEW P6 XER schedule file (the actual state of the schedule)\n\n"
            "For each comment where the contractor claims 'addressed' or 'resolved' but the XER evidence "
            "contradicts that claim, produce an exception entry.\n\n"
            "Output ONLY a valid JSON array of objects. Each object must have these keys:\n"
            "- 'Exception ID': Sequential EXC-001, EXC-002, etc.\n"
            "- 'Original Comment ID': The BLR-NNN ID from the prior review that the contractor claimed to address.\n"
            "- 'WBS Reference': The WBS element or activity ID involved.\n"
            "- 'Original Comment Description': Brief summary of what the original comment required.\n"
            "- 'Spec Section Reference': The contract spec section being violated.\n"
            "- 'Contractor Response': What the contractor said in their response (summarized).\n"
            "- 'Contractor Claimed Addressed': Always 'Yes' (these are only the ones contractor claimed fixed).\n"
            "- 'Agent Evaluation': 'NOT ADDRESSED' for all entries in this report.\n"
            "- 'Evidence from XER': Specific evidence from the XER file showing the issue persists "
            "(e.g., 'Activity X still has 0-day duration', 'Logic tie still missing between A and B').\n"
            "- 'Spec Non-Compliance Detail': How this violates the contract specification or CPM best practice.\n"
            "- 'Severity': 'Critical' (spec violation / logic error) or 'Major' (best-practice non-compliance).\n\n"
            "Rules:\n"
            "- ONLY include entries where contractor claimed 'addressed/yes/resolved' but the XER proves otherwise.\n"
            "- Do NOT include entries where the contractor genuinely fixed the issue.\n"
            "- Do NOT include entries where the contractor marked 'No' or left blank.\n"
            "- If there are NO exceptions (contractor addressed everything properly), return an empty array [].\n"
            "- Be specific with XER evidence — cite activity IDs, durations, logic ties, float values.\n"
            "- Pay special attention to logic-related comments (Missing Hard Logic, Preferential Cross-WBS, "
            "Dangling Logic). If the contractor claimed these were addressed but the XER TASKPRED table "
            "still shows the logic deficiency, flag as exception with specific predecessor/successor evidence.\n"
            "- Output ONLY the JSON array, no markdown, no explanation."
        )

        exc_user_parts = []
        if specs_content:
            exc_user_parts.append(f"## Contract Specifications\n{specs_content[:12000]}")
        exc_user_parts.append(
            f"## Contractor's Filled Comment Response (Previous Version)\n{resp_content[:12000]}"
        )
        exc_user_parts.append(
            f"## New P6 XER Schedule ({review_label} v{ver})\n{_summarize_xer_for_review(xer_content)[:80000]}"
        )
        exc_user_msg = "\n\n".join(exc_user_parts)
        exc_user_msg += (
            "\n\nIdentify all exceptions where the contractor claimed 'addressed' "
            "but the XER evidence shows otherwise. Output as JSON array."
        )

        exc_reply, exc_err = _call_llm(exc_system, [{"role": "user", "content": exc_user_msg}])
        if not exc_err and exc_reply:
            exc_raw = exc_reply.strip()
            if exc_raw.startswith("```"):
                exc_raw = exc_raw.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
            try:
                exceptions = json.loads(exc_raw)
                if isinstance(exceptions, list) and len(exceptions) > 0:
                    for idx, e in enumerate(exceptions):
                        e["Exception ID"] = f"EXC-{idx + 1:03d}"
                    exception_filepath = save_exception_report(request.session_id, ver, exceptions)
                    exception_count = len(exceptions)
            except Exception:
                pass

    meta = save_review_meta(
        request.session_id, ver, columns, len(comments), filepath, last_comment_num,
        exception_count=exception_count, exception_filepath=exception_filepath,
    )

    try:
        from schedule_agent_web.vector_store import index_file
        review_text = "\n".join(
            f"[{c.get('Comment ID','')}] {c.get('Priority','')} | {c.get('Logic Flag','')} | "
            f"WBS: {c.get('WBS Reference','')} | {c.get(next((k for k in c if k.startswith('Comment Description')), ''),'')} | "
            f"Spec: {c.get('Spec Section Reference','')} | "
            f"Rec: {c.get(next((k for k in c if k.startswith('Recommendation')), ''), '')}"
            for c in comments
        )
        index_file(request.session_id, f"_baseline_review_v{ver}", review_text)
        if exception_count > 0 and exceptions:
            exc_text = "\n".join(
                f"[{e.get('Exception ID','')}] {e.get('Original Comment ID','')} | "
                f"WBS: {e.get('WBS Reference','')} | {e.get('Evidence from XER','')} | "
                f"Severity: {e.get('Severity','')}"
                for e in exceptions
            )
            index_file(request.session_id, f"_baseline_exception_v{ver}", exc_text)
    except Exception:
        pass

    return {"ok": True, "review": meta, "comments": comments}


@app.get("/api/baseline/reviews")
def api_baseline_reviews(session_id: str = ""):
    """List all review results for a project."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    from schedule_agent_web.baseline import list_reviews
    return {"reviews": list_reviews(session_id)}


@app.get("/api/baseline/review/download")
def api_baseline_review_download(session_id: str = "", version: int = 0):
    """Download review comments Excel file."""
    if not session_id or not version:
        raise HTTPException(status_code=400, detail="session_id and version required")
    from schedule_agent_web.baseline import get_review
    review = get_review(session_id, version)
    if not review:
        raise HTTPException(status_code=404, detail=f"No review found for v{version}")
    filepath = review["filepath"]
    if not os.path.isfile(filepath):
        raise HTTPException(status_code=404, detail="Review file not found on disk.")
    fname = os.path.basename(filepath)
    return FileResponse(filepath, filename=fname, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/baseline/review/view")
def api_baseline_review_view(session_id: str = "", version: int = 0):
    """Return review comments as JSON for on-screen viewing."""
    if not session_id or not version:
        raise HTTPException(status_code=400, detail="session_id and version required")
    from schedule_agent_web.baseline import get_review
    review = get_review(session_id, version)
    if not review:
        raise HTTPException(status_code=404, detail=f"No review found for v{version}")
    filepath = review["filepath"]
    if not os.path.isfile(filepath):
        raise HTTPException(status_code=404, detail="Review file not found on disk.")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return {"columns": [], "comments": []}
        columns = [str(c) if c else "" for c in rows[0]]
        comments = []
        for row in rows[1:]:
            comment = {}
            for i, col in enumerate(columns):
                val = row[i] if i < len(row) else None
                comment[col] = val if val is not None else ""
            comments.append(comment)
        return {"columns": columns, "comments": comments, "version": version}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read review file: {e}")


@app.get("/api/baseline/review/exception/download")
def api_baseline_exception_download(session_id: str = "", version: int = 0):
    """Download PCM Exception Report Excel file."""
    if not session_id or not version:
        raise HTTPException(status_code=400, detail="session_id and version required")
    from schedule_agent_web.baseline import get_review
    review = get_review(session_id, version)
    if not review:
        raise HTTPException(status_code=404, detail=f"No review found for v{version}")
    filepath = review.get("exception_filepath", "")
    if not filepath or not os.path.isfile(filepath):
        raise HTTPException(status_code=404, detail="No exception report available for this version.")
    fname = os.path.basename(filepath)
    return FileResponse(filepath, filename=fname, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.delete("/api/baseline/review")
def api_baseline_review_delete(session_id: str = "", version: int = 0):
    """Delete a baseline review result so the user can re-run the review."""
    if not session_id or not version:
        raise HTTPException(status_code=400, detail="session_id and version required")
    from schedule_agent_web.baseline import delete_review
    deleted = delete_review(session_id, version)
    if not deleted:
        raise HTTPException(status_code=404, detail=f"No review found for v{version}")
    return {"ok": True, "message": f"Review for v{version} deleted. You can now re-run the review."}


# --- Claude with vision (site images / project progress) ---
@app.post("/api/vision/describe")
async def api_vision_describe(
    file: UploadFile = File(...),
    prompt: str = Form(""),
):
    """
    Describe an image using Claude with vision (e.g. site photo for project progress).
    Requires ANTHROPIC_API_KEY. Accepts JPEG, PNG, GIF, WebP.
    Returns { description, error }. Optional prompt: custom instruction (default asks for site progress summary).
    """
    default_prompt = (
        "Describe what you see in this image. If it looks like a construction or project site, "
        "summarize the visible progress, work in place, and any notable conditions (e.g. scaffolding, concrete, MEP, safety). Be concise."
    )
    try:
        raw = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")
    if len(raw) > 500 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image too large (max 500MB)")
    media_type = (file.content_type or "image/jpeg").strip().lower()
    if media_type not in ("image/jpeg", "image/png", "image/gif", "image/webp"):
        media_type = "image/jpeg"
    user_prompt = (prompt or "").strip() or default_prompt
    try:
        from schedule_agent_web.vision import describe_image
        description, err = describe_image(raw, media_type=media_type, prompt=user_prompt)
        if err:
            raise HTTPException(status_code=502, detail=err)
        return {"description": description, "error": None}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/vision/ingest")
async def api_vision_ingest(
    file: UploadFile = File(...),
    session_id: str = Form(""),
    group: str = Form("contract_specs"),
):
    """
    Ingest an image via the same pipeline as POST /api/ingest/document. Vision runs inside the ingestion layer:
    extractors call Claude to describe the image, then the description is stored and vectorized.
    Requires ANTHROPIC_API_KEY. Use this endpoint for image-only uploads; or upload images to /api/ingest/document.
    """
    if not (session_id or "").strip():
        raise HTTPException(status_code=400, detail="session_id required")
    session_id = session_id.strip()
    try:
        raw = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")
    if len(raw) > 500 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Image too large (max 500MB)")
    filename = file.filename or "site-photo.jpg"
    if not filename.lower().endswith((".jpg", ".jpeg", ".png", ".gif", ".webp")):
        filename = filename + ".jpg"
    try:
        from schedule_agent_web.ingestion import ingest_document, get_ingested_document
        ingestion_group = (group or "contract_specs").strip() or "contract_specs"
        result = ingest_document(session_id, filename, raw, source="vision", ingestion_group=ingestion_group)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    if result.get("status") == "error":
        raise HTTPException(status_code=500, detail=result.get("error", "Ingestion failed"))
    if result.get("status") == "created" and result.get("doc_id"):
        try:
            doc = get_ingested_document(session_id, result["doc_id"])
            if doc and doc.raw_text:
                from schedule_agent_web.store import save_file
                save_file(session_id, filename, doc.raw_text)
                from schedule_agent_web.vector_store import index_file
                index_file(session_id, filename, doc.raw_text, ingestion_group=ingestion_group)
        except Exception:
            pass
    return result


# --- Stage 1: Ingestion & Normalization (NLP Doc Intel Pipeline) ---
@app.get("/api/ingest/groups")
def api_ingest_groups():
    """Return ingestion groups for the ingestion page (contract docs vs sample schedule)."""
    try:
        from schedule_agent_web.ingestion.pipeline import INGESTION_GROUPS
        return INGESTION_GROUPS
    except Exception:
        return [
            {"id": "contract_specs", "label": "Contract documents & specifications"},
            {"id": "sample_schedule", "label": "Sample schedule (reference for building new schedule)"},
            {"id": "site_progress", "label": "Site pictures & daily logs"},
        ]


@app.post("/api/ingest/document")
async def api_ingest_document(
    file: UploadFile = File(...),
    session_id: str = Form(""),
    group: str = Form("contract_specs"),
):
    """
    Ingest a document into the normalized store (PDF, DOCX, XLSX, TXT, CSV, XER, or site images JPEG/PNG/GIF/WebP).
    Images are described via Claude vision (ANTHROPIC_API_KEY) and the description is stored and vectorized.
    group: "contract_specs" | "sample_schedule" for ingestion page sections.
    Extracts text (or image description), metadata, content hash; deduplicates by group; saves and vectorizes for search.
    """
    if not (session_id or "").strip():
        raise HTTPException(status_code=400, detail="session_id required")
    session_id = session_id.strip()
    filename = file.filename or "unknown"
    ingestion_group = (group or "contract_specs").strip() or "contract_specs"
    try:
        raw = await file.read()
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read file: {e}")
    if len(raw) > 500 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 500MB)")
    try:
        from schedule_agent_web.ingestion import ingest_document
        result = ingest_document(session_id, filename, raw, source="upload", ingestion_group=ingestion_group)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    if result.get("status") == "error":
        raise HTTPException(status_code=500, detail=result.get("error", "Ingestion failed"))
    # Backfill file store and vectorize so chat/RAG and search can use it (with group for filtered search)
    if result.get("status") == "created" and result.get("doc_id"):
        try:
            from schedule_agent_web.ingestion import get_ingested_document
            doc = get_ingested_document(session_id, result["doc_id"])
            if doc and doc.raw_text:
                from schedule_agent_web.store import save_file, update_file_meta
                save_file(session_id, filename, doc.raw_text)
                vectorized = False
                try:
                    from schedule_agent_web.vector_store import index_file
                    vectorized = index_file(session_id, filename, doc.raw_text, ingestion_group=ingestion_group)
                except Exception:
                    pass
                group_to_cat = {
                    "contract_specs": "Contract Specifications",
                    "sample_schedule": "Schedule Documents",
                    "site_progress": "Photos / Site Progress",
                }
                cat = group_to_cat.get(ingestion_group, _auto_category_from_filename(filename))
                update_file_meta(session_id, filename, vectorized=vectorized, category=cat)
        except Exception:
            pass
    return result


@app.get("/api/ingest/documents")
def api_list_ingested_documents(session_id: str = "", group: str = ""):
    """List documents in the Stage 1 normalized store. Optional group= filters to that ingestion group."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.ingestion import list_ingested_documents
        return list_ingested_documents(session_id, ingestion_group=group.strip() or None)
    except Exception:
        return []


@app.get("/api/ingest/documents/{doc_id}")
def api_get_ingested_document(doc_id: str, session_id: str = ""):
    """Return one normalized document (id, filename, format, raw_text, metadata, content_hash, created_at)."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    try:
        from schedule_agent_web.ingestion import get_ingested_document
        doc = get_ingested_document(session_id, doc_id)
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        return doc.to_dict()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- Stage 2: Preprocessing & Enrichment (NLP Doc Intel Pipeline) ---
class EnrichDocumentRequest(BaseModel):
    session_id: str
    doc_id: str


@app.post("/api/enrich/document")
def api_enrich_document(request: EnrichDocumentRequest):
    """Run Stage 2 (cleaning, language, segmentation, structure, vocabulary) on an ingested doc. Requires doc_id from Stage 1."""
    if not request.session_id or not request.doc_id:
        raise HTTPException(status_code=400, detail="session_id and doc_id required")
    try:
        from schedule_agent_web.ingestion import get_ingested_document
        from schedule_agent_web.enrichment import enrich_document
        doc = get_ingested_document(request.session_id, request.doc_id)
        if not doc:
            raise HTTPException(status_code=404, detail="Ingested document not found. Ingest the document first (Stage 1).")
        result = enrich_document(request.session_id, request.doc_id, doc.raw_text)
        if result.get("status") == "error":
            raise HTTPException(status_code=500, detail=result.get("error", "Enrichment failed"))
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/enrich/documents")
def api_list_enriched_documents(session_id: str = ""):
    """List Stage 2 enriched documents for this session."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.enrichment import list_enriched_documents
        return list_enriched_documents(session_id)
    except Exception:
        return []


@app.get("/api/enrich/documents/{doc_id}")
def api_get_enriched_document(doc_id: str, session_id: str = ""):
    """Return one enriched document (cleaned_text, lang, sentences, structure, normalized_text, term_replacements)."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    try:
        from schedule_agent_web.enrichment import get_enriched_document
        doc = get_enriched_document(session_id, doc_id)
        if not doc:
            raise HTTPException(status_code=404, detail="Enriched document not found")
        return doc.to_dict()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- Stage 3: Core NLP (NER, classification, relations, temporal, summarization) ---
class ProcessNLPRequest(BaseModel):
    session_id: str
    doc_id: str
    use_spacy_ner: bool = False


@app.post("/api/nlp/process")
def api_nlp_process(request: ProcessNLPRequest):
    """Run Stage 3 NLP on an ingested doc. Uses raw text from Stage 1 (or enriched from Stage 2). Requires OPENAI_API_KEY or ANTHROPIC_API_KEY for classification, relations, summary."""
    if not request.session_id or not request.doc_id:
        raise HTTPException(status_code=400, detail="session_id and doc_id required")
    try:
        from schedule_agent_web.ingestion import get_ingested_document
        from schedule_agent_web.enrichment import get_enriched_document
        from schedule_agent_web.nlp import process_document
        # Prefer enriched text if available; else raw from Stage 1
        text = None
        enriched = get_enriched_document(request.session_id, request.doc_id)
        if enriched and enriched.cleaned_text:
            text = enriched.normalized_text or enriched.cleaned_text
        if not text:
            doc = get_ingested_document(request.session_id, request.doc_id)
            if not doc:
                raise HTTPException(status_code=404, detail="Document not found. Ingest (Stage 1) first.")
            text = doc.raw_text
        result = process_document(
            request.session_id,
            request.doc_id,
            text,
            use_spacy_ner=request.use_spacy_ner,
        )
        if result.get("status") == "error":
            raise HTTPException(status_code=500, detail=result.get("error", "NLP processing failed"))
        # Stage 4: save relations to knowledge graph; record analytics events
        try:
            from schedule_agent_web.nlp import get_nlp_document
            from schedule_agent_web.intelligence import save_edges, record_event
            nlp_doc = get_nlp_document(request.session_id, request.doc_id)
            if nlp_doc:
                save_edges(request.session_id, request.doc_id, nlp_doc.relations)
                record_event(request.session_id, request.doc_id, "nlp_processed", {"summary_preview": (nlp_doc.summary or "")[:200]})
                cl = nlp_doc.classification or {}
                if cl.get("risk_signal"):
                    record_event(request.session_id, request.doc_id, "risk_signal")
                if cl.get("change_signal"):
                    record_event(request.session_id, request.doc_id, "change_signal")
        except Exception:
            pass
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/nlp/documents")
def api_list_nlp_documents(session_id: str = ""):
    """List Stage 3 NLP-processed documents for this session."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.nlp import list_nlp_documents
        return list_nlp_documents(session_id)
    except Exception:
        return []


@app.get("/api/nlp/documents/{doc_id}")
def api_get_nlp_document(doc_id: str, session_id: str = ""):
    """Return one NLP document (entities, relations, classification, dates, summary)."""
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    try:
        from schedule_agent_web.nlp import get_nlp_document
        doc = get_nlp_document(session_id, doc_id)
        if not doc:
            raise HTTPException(status_code=404, detail="NLP document not found")
        return doc.to_dict()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# --- Stage 4: Signals, knowledge graph, analytics ---
class ScanSignalsRequest(BaseModel):
    session_id: str
    doc_id: str
    use_llm: bool = False


@app.post("/api/signals/scan")
def api_signals_scan(request: ScanSignalsRequest):
    """Run signal detection (rules + optional LLM) on an ingested doc's text. Saves signals to store."""
    if not request.session_id or not request.doc_id:
        raise HTTPException(status_code=400, detail="session_id and doc_id required")
    try:
        from schedule_agent_web.ingestion import get_ingested_document
        from schedule_agent_web.enrichment import get_enriched_document
        from schedule_agent_web.intelligence import scan_document_for_signals
        text = None
        enriched = get_enriched_document(request.session_id, request.doc_id)
        if enriched and enriched.cleaned_text:
            text = enriched.cleaned_text
        if not text:
            doc = get_ingested_document(request.session_id, request.doc_id)
            if not doc:
                raise HTTPException(status_code=404, detail="Document not found. Ingest first.")
            text = doc.raw_text
        signals = scan_document_for_signals(
            request.session_id, request.doc_id, text,
            use_llm=request.use_llm,
        )
        # Record events for trend analytics
        try:
            from schedule_agent_web.intelligence import record_event
            for s in signals:
                t = (s.get("signal_type") or "risk") + "_signal"
                record_event(request.session_id, request.doc_id, t)
        except Exception:
            pass
        # Stage 5: audit trail + alerts for high-priority signals
        try:
            from schedule_agent_web.delivery import append_audit, send_alert
            for s in signals:
                append_audit(request.session_id, "signal", {
                    "doc_id": request.doc_id,
                    "signal_type": s.get("signal_type", "risk"),
                    "source": s.get("source", "rule"),
                    "snippet": (s.get("text_snippet") or s.get("snippet") or "")[:300],
                })
                st = s.get("signal_type") or "risk"
                if st in ("risk", "dispute"):
                    send_alert(request.session_id, st, request.doc_id, s.get("text_snippet") or s.get("snippet", ""), s.get("source", "rule"))
        except Exception:
            pass
        return {"status": "ok", "count": len(signals), "signals": signals}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/signals")
def api_list_signals(session_id: str = "", signal_type: str = ""):
    """List detected signals for this session. Optional signal_type: risk, change, dispute."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.intelligence import list_signals
        return list_signals(session_id, signal_type=signal_type or None)
    except Exception:
        return []


@app.get("/api/graph/related")
def api_graph_related(entity: str = "", session_id: str = "", limit: int = 50):
    """Return knowledge graph edges (subject, relation, object) where subject or object matches entity."""
    if not session_id or not (entity or "").strip():
        return []
    try:
        from schedule_agent_web.intelligence import get_related
        return get_related(session_id, entity.strip(), limit=max(1, min(limit, 100)))
    except Exception:
        return []


@app.get("/api/analytics/trends")
def api_analytics_trends(session_id: str = "", from_date: str = "", to_date: str = "", bucket_days: int = 7):
    """Trend aggregates: event counts per time bucket (nlp_processed, risk_signal, change_signal, dispute_signal)."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.intelligence import get_trends
        return get_trends(
            session_id,
            from_date=from_date or None,
            to_date=to_date or None,
            bucket_days=max(1, min(bucket_days, 90)),
        )
    except Exception:
        return []


# --- Stage 5: Delivery & integration ---
@app.get("/api/dashboard")
def api_dashboard(session_id: str = "", recent_limit: int = 10):
    """Dashboard payload for UI: signals count, recent signals, NLP summaries, ingested count, trends."""
    if not session_id:
        return {"signals_count": 0, "recent_signals": [], "nlp_summaries": [], "ingested_count": 0, "trends": []}
    try:
        from schedule_agent_web.delivery import get_dashboard
        return get_dashboard(session_id, recent_limit=max(1, min(recent_limit, 50)))
    except Exception:
        return {"signals_count": 0, "recent_signals": [], "nlp_summaries": [], "ingested_count": 0, "trends": []}


class DocumentsQARequest(BaseModel):
    session_id: str
    question: str


@app.post("/api/documents/qa")
def api_documents_qa(request: DocumentsQARequest):
    """Standalone Q&A over session documents (RAG + LLM). Returns answer and sources. Records audit."""
    if not request.session_id or not (request.question or "").strip():
        raise HTTPException(status_code=400, detail="session_id and question required")
    if not get_status_dict().get("has_api_key"):
        raise HTTPException(status_code=503, detail="Set OPENAI_API_KEY or ANTHROPIC_API_KEY.")
    try:
        from schedule_agent_web.vector_store import search, is_qdrant_available, _file_search_fallback
        from schedule_agent_web.nlp.llm_utils import call_llm
        from schedule_agent_web.delivery import append_audit
        if is_qdrant_available():
            hits = search(request.session_id, request.question.strip(), top_k=6)
        else:
            hits = _file_search_fallback(request.session_id, request.question.strip(), top_k=6)
        context = ""
        sources = []
        if hits:
            for h in hits:
                fn = h.get("filename", "")
                txt = (h.get("text") or "")[:800]
                context += f"\nFrom {fn}:\n{txt}\n"
                sources.append({"filename": fn, "text_preview": txt[:300]})
        system = "You are a project controls analyst. Answer the question using only the provided context. If the context does not contain enough information, say so. Be concise."
        reply, err = call_llm(system, f"Context:\n{context}\n\nQuestion: {request.question.strip()}", max_tokens=1024)
        if err:
            raise HTTPException(status_code=502, detail=err)
        append_audit(request.session_id, "answer", {"source": "documents_qa", "sources": sources})
        return {"answer": reply, "sources": sources}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class ReportDraftRequest(BaseModel):
    session_id: str
    title: str
    template: str = ""


@app.post("/api/reports/draft")
def api_reports_draft(request: ReportDraftRequest):
    """Generate a report draft from NLP summaries and signals (LLM). Records audit."""
    if not request.session_id or not (request.title or "").strip():
        raise HTTPException(status_code=400, detail="session_id and title required")
    if not get_status_dict().get("has_api_key"):
        raise HTTPException(status_code=503, detail="Set OPENAI_API_KEY or ANTHROPIC_API_KEY.")
    try:
        from schedule_agent_web.delivery import draft_report, append_audit
        draft, err = draft_report(request.session_id, request.title.strip(), template_hint=request.template.strip())
        if err:
            raise HTTPException(status_code=502, detail=err)
        append_audit(request.session_id, "report", {"title": request.title.strip(), "source": "reports_draft"})
        return {"draft": draft, "title": request.title.strip()}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/audit/trail")
def api_audit_trail(session_id: str = "", limit: int = 100, entry_type: str = ""):
    """Audit trail: answers, signals, reports (traceable to source). Optional entry_type=answer|signal|report."""
    if not session_id:
        return []
    try:
        from schedule_agent_web.delivery import get_audit_trail
        return get_audit_trail(session_id, limit=max(1, min(limit, 500)), entry_type=entry_type or None)
    except Exception:
        return []


@app.delete("/api/files")
def api_delete_file(session_id: str = "", filename: str = ""):
    """Delete one uploaded file (cannot delete internal/library files)."""
    if not session_id or not filename:
        raise HTTPException(status_code=400, detail="session_id and filename required")
    if filename.startswith("_"):
        raise HTTPException(status_code=403, detail="Cannot delete internal library files.")
    try:
        from schedule_agent_web.store import delete_file
        if delete_file(session_id, filename):
            try:
                from schedule_agent_web.vector_store import delete_file_vectors
                delete_file_vectors(session_id, filename)
            except Exception:
                pass
            try:
                from schedule_agent_web.baseline import delete_submissions_by_file
                delete_submissions_by_file(session_id, filename)
            except Exception:
                pass
            return {"status": "ok"}
        raise HTTPException(status_code=404, detail="File not found")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _parse_json_from_llm(reply: str) -> dict | None:
    """Extract a JSON object from LLM reply (markdown code block, raw JSON, or first {...})."""
    import re
    import json as _json
    reply = (reply or "").strip().lstrip("\ufeff")
    if not reply:
        return None

    def try_load(s: str) -> dict | None:
        s = s.strip()
        if not s:
            return None
        # Remove trailing commas before ] or } (invalid in JSON but some models output them)
        s = re.sub(r",\s*([}\]])", r"\1", s)
        try:
            return _json.loads(s)
        except Exception:
            pass
        return None

    # 1) Try ```json ... ``` or ``` ... ``` (take first code block)
    m = re.search(r"```(?:json)?\s*([\s\S]*?)```", reply)
    if m:
        out = try_load(m.group(1))
        if out:
            return out
    # 2) Try first line that looks like raw JSON (starts with {)
    for line in reply.split("\n"):
        line = line.strip()
        if line.startswith("{"):
            out = try_load(line)
            if out:
                return out
    # 3) Find outermost {...} (brace match)
    start = reply.find("{")
    if start >= 0:
        depth = 0
        in_string = None
        escape = False
        i = start
        while i < len(reply):
            c = reply[i]
            if escape:
                escape = False
                i += 1
                continue
            if c == "\\" and in_string:
                escape = True
                i += 1
                continue
            if in_string:
                if c == in_string:
                    in_string = None
                i += 1
                continue
            if c in ('"', "'"):
                in_string = c
                i += 1
                continue
            if c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    out = try_load(reply[start : i + 1])
                    if out:
                        return out
                    break
            i += 1
    # 4) Fallback: first { to last } (handles trailing explanation or extra text after JSON)
    if start >= 0:
        last_brace = reply.rfind("}")
        if last_brace > start:
            out = try_load(reply[start : last_brace + 1])
            if out:
                return out
    return None


def _extract_json_array(reply: str) -> list | None:
    """Extract a JSON array from LLM reply. Handles markdown fences, extra text, trailing commas."""
    import re as _re
    raw = (reply or "").strip()
    if not raw:
        return None

    def _clean_and_parse(s):
        s = s.strip()
        if not s:
            return None
        s = _re.sub(r",\s*([}\]])", r"\1", s)
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return parsed
        except Exception:
            pass
        return None

    m = _re.search(r"```(?:json)?\s*([\s\S]*?)```", raw)
    if m:
        result = _clean_and_parse(m.group(1))
        if result is not None:
            return result

    result = _clean_and_parse(raw)
    if result is not None:
        return result

    start = raw.find("[")
    if start >= 0:
        last_bracket = raw.rfind("]")
        if last_bracket > start:
            result = _clean_and_parse(raw[start:last_bracket + 1])
            if result is not None:
                return result

    return None


def _normalize_activity(a: dict) -> dict:
    """Ensure activity has all fields required by frontend Activity type."""
    preds = a.get("predecessors") or []
    return {
        "id": str(a.get("id") or f"ACT-{hash(a.get('name', '')) % 100000}"),
        "name": str(a.get("name") or "Unnamed"),
        "durationDays": int(a.get("durationDays") or a.get("duration") or 0),
        "totalFloatDays": int(a.get("totalFloatDays") or a.get("float") or 0),
        "wbsCode": str(a.get("wbsCode") or a.get("wbs") or ""),
        "predecessors": [
            {
                "activityId": str(p.get("activityId") or p.get("id") or ""),
                "type": str(p.get("type") or "FS")[:2],
                "lag": int(p.get("lag") or 0),
            }
            for p in preds if isinstance(p, dict)
        ],
        "isCritical": bool(a.get("isCritical", False)),
        "healthFlags": list(a.get("healthFlags") or []),
        "directCost": int(a.get("directCost") or 0),
        "indirectCost": int(a.get("indirectCost") or 0),
        "isHardCost": bool(a.get("isHardCost", True)),
    }


class ExtractScheduleRequest(BaseModel):
    session_id: str


@app.get("/api/schedule")
def api_get_schedule(session_id: str = ""):
    """Return saved extracted schedule (activities, wbs, projectName) for this session."""
    if not session_id:
        return None
    try:
        from schedule_agent_web.store import get_schedule
        return get_schedule(session_id) or None
    except Exception:
        return None


@app.post("/api/extract-schedule")
def api_extract_schedule(request: ExtractScheduleRequest):
    """Extract activities and WBS from uploaded scope documents; save and return for Activity List."""
    if not request.session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    if not get_status_dict().get("has_api_key"):
        raise HTTPException(status_code=503, detail="Set OPENAI_API_KEY or ANTHROPIC_API_KEY in environment.")
    from schedule_agent_web.store import get_files
    files = [f for f in get_files(request.session_id) if not f.get("filename", "").startswith("_")]
    if not files:
        try:
            from schedule_agent_web.ingestion import list_ingested_documents, get_ingested_document
            ingested = list_ingested_documents(request.session_id, ingestion_group="contract_specs")
            if not ingested:
                raise HTTPException(
                    status_code=400,
                    detail="No scope documents found. In the Ingestion tab, upload scope files in 'Scope documents' (PDF, TXT, etc.), then try again.",
                )
            # Inject ingested content into context via a one-off system addition
            files = [{"filename": d.get("filename", "doc"), "source": "ingested"} for d in ingested]
        except HTTPException:
            raise
        except Exception:
            raise HTTPException(
                status_code=400,
                detail="No scope documents uploaded. Upload PDF or text scope files in the Ingestion tab (Scope documents), then try again.",
            )
    system = get_system_prompt_with_context(request.session_id)
    # If we only have ingested docs, append their raw text (get_system_prompt_with_context uses get_files)
    if files and files[0].get("source") == "ingested":
        try:
            from schedule_agent_web.ingestion import list_ingested_documents, get_ingested_document
            ingested = list_ingested_documents(request.session_id, ingestion_group="contract_specs")
            system += "\n\n## Uploaded project files (from Ingestion)\n"
            for d in ingested[:20]:
                doc_id = d.get("id") or d.get("doc_id")
                if not doc_id:
                    continue
                doc = get_ingested_document(request.session_id, doc_id)
                if doc and doc.raw_text:
                    preview = doc.raw_text[:80000] if len(doc.raw_text) > 80000 else doc.raw_text
                    system += f"\n### File: {doc.filename}\n```\n{preview}\n```\n"
        except Exception:
            pass
    system += "\n\nCRITICAL: You must respond with exactly one valid JSON object. No markdown, no code fences, no text before or after. Start your response with { and end with }. Use double quotes for strings. No trailing commas."
    system += '\nRequired JSON shape: {"projectName": "string", "activities": [{"id": "ACT-001", "name": "string", "durationDays": 0, "wbsCode": "1.1", "predecessors": [{"activityId": "ACT-000", "type": "FS", "lag": 0}], "isCritical": false, "totalFloatDays": 0, "healthFlags": [], "directCost": 0, "indirectCost": 0, "isHardCost": true}], "wbs": [{"code": "1", "name": "Phase name"}]}'
    user_msg = (
        "From the uploaded project files above, extract a CPM schedule. "
        "List every activity with: id (e.g. ACT-001), name, durationDays (number), wbsCode (e.g. 1.1), predecessors (list of {activityId, type: FS|SS|FF|SF, lag: number}), isCritical (boolean), totalFloatDays (0 if unknown), healthFlags ([]), directCost (0 if unknown), indirectCost (0), isHardCost (true). "
        "Include wbs as a list of {code, name}. Set projectName from the scope. "
        "Reply with ONLY the JSON object—no explanation, no markdown code block, no backticks. Start with { and end with }."
    )
    reply, err = _call_llm(system, [{"role": "user", "content": user_msg}])
    if err:
        raise HTTPException(status_code=502, detail=err)
    data = _parse_json_from_llm(reply)
    # Retry once with a minimal "JSON only" prompt if first response didn't parse
    if not data and reply:
        retry_system = system + "\n\nYou must reply with ONLY a single valid JSON object. No markdown, no code block, no text before or after. Start with { and end with }."
        retry_msg = (
            "Output ONLY valid JSON. No explanation. Use this exact structure; add as many activities as the scope describes: "
            '{"projectName": "string", "activities": [{"id": "ACT-001", "name": "string", "durationDays": 0, "wbsCode": "1.1", "predecessors": [], "isCritical": false, "totalFloatDays": 0, "healthFlags": [], "directCost": 0, "indirectCost": 0, "isHardCost": true}], "wbs": [{"code": "1", "name": "string"}]}'
        )
        reply2, err2 = _call_llm(retry_system, [{"role": "user", "content": retry_msg}])
        if not err2:
            data = _parse_json_from_llm(reply2)
        if not data and reply2:
            reply = reply2  # use retry reply for snippet below
    if not data:
        snippet = (reply or "")[:400].replace("\n", " ").strip()
        if snippet:
            detail = f"The model did not return valid JSON. Model reply (first 400 chars): {snippet} …"
        else:
            detail = "The model returned an empty response. Try a shorter scope document (1–2 pages) and click Extract again."
        raise HTTPException(status_code=502, detail=detail)
    if not isinstance(data.get("activities"), list):
        data["activities"] = []
    activities = [_normalize_activity(a) for a in data["activities"] if isinstance(a, dict)]
    wbs = data.get("wbs") or []
    if not isinstance(wbs, list):
        wbs = []
    project_name = data.get("projectName") or "Extracted from scope"
    out = {"projectName": project_name, "activities": activities, "wbs": wbs}
    try:
        from schedule_agent_web.store import save_schedule
        save_schedule(request.session_id, out)
    except Exception:
        pass
    return out


def _call_llm(system: str, messages_for_llm: list, max_tokens: int = 8192) -> tuple[str, str | None]:
    """Call OpenAI or Claude; returns (reply, error). Prefers Claude if ANTHROPIC_API_KEY set."""
    if _use_claude():
        key = _get_anthropic_key()
        default_models = ("claude-sonnet-4-20250514", "claude-3-5-sonnet-20241022", "claude-3-5-haiku-20241022", "claude-3-haiku-20240307")
        model = os.environ.get("ANTHROPIC_MODEL") or default_models[0]
        models_to_try = [model] + [m for m in default_models if m != model]
        claude_messages = [{"role": m["role"], "content": m["content"]} for m in messages_for_llm if m["role"] in ("user", "assistant")]
        if max_tokens > 64000:
            max_tokens = 64000
        last_err = None
        for try_model in models_to_try:
            try:
                client = Anthropic(api_key=key)
                if max_tokens > 8192:
                    collected = []
                    with client.messages.stream(
                        model=try_model,
                        max_tokens=max_tokens,
                        system=system,
                        messages=claude_messages,
                    ) as stream:
                        for text_chunk in stream.text_stream:
                            collected.append(text_chunk)
                    text = "".join(collected).strip()
                else:
                    resp = client.messages.create(
                        model=try_model,
                        max_tokens=max_tokens,
                        system=system,
                        messages=claude_messages,
                    )
                    text = (resp.content[0].text if resp.content else "").strip()
                return (text, None)
            except Exception as e:
                last_err = e
                if "404" in str(e) or "not_found" in str(e).lower():
                    continue
                return ("", str(e))
        return ("", str(last_err) if last_err else "No Claude model available. Set ANTHROPIC_MODEL in Vercel to a model your account has (e.g. claude-3-haiku-20240307).")
    # OpenAI
    key = _get_openai_key()
    if not key or not OpenAI:
        return ("", "No API key or OpenAI not installed. Set OPENAI_API_KEY or ANTHROPIC_API_KEY.")
    try:
        client = OpenAI(api_key=key)
        resp = client.chat.completions.create(
            model=os.environ.get("OPENAI_CHAT_MODEL", "gpt-4o-mini"),
            messages=[{"role": "system", "content": system}] + messages_for_llm,
            temperature=0.3,
            max_tokens=max_tokens,
        )
        reply = (resp.choices[0].message.content or "").strip()
        return (reply, None)
    except Exception as e:
        err = str(e)
        if "429" in err or "quota" in err.lower() or "insufficient_quota" in err.lower():
            err += " Use Claude instead: in Vercel set ANTHROPIC_API_KEY (get a key at console.anthropic.com) and redeploy."
        return ("", err)


@app.post("/api/chat", response_model=ChatResponse)
def chat(request: ChatRequest):
    """Send a message to the Schedule Agent (VueLogic). All conversations are persisted and used as context (agent learns from history)."""
    if not request.message.strip():
        raise HTTPException(status_code=400, detail="message is required")
    if not get_status_dict().get("has_api_key"):
        raise HTTPException(
            status_code=503,
            detail="Set OPENAI_API_KEY or ANTHROPIC_API_KEY in environment.",
        )
    system = get_system_prompt_with_context(request.session_id)
    # RAG: retrieve from Qdrant (+ optional GraphRAG) and inject into context; collect sources for citation (Stage 4)
    rag_sources = []
    try:
        from schedule_agent_web.vector_store import search, graphrag_search, is_qdrant_available, is_graphrag_available, _file_search_fallback
        rag_parts = []
        import logging
        _chat_log = logging.getLogger("chat_rag")
        if is_qdrant_available():
            hits = search(request.session_id, request.message.strip(), top_k=8)
            _chat_log.info("RAG via Qdrant: %d hits", len(hits))
        else:
            hits = _file_search_fallback(request.session_id, request.message.strip(), top_k=8)
            _chat_log.info("RAG via fallback: %d hits", len(hits))
        if hits:
            doc_filenames = set()
            rag_parts.append("\n\n## Retrieved from uploaded documents and past conversations (use when answering)\n")
            for h in hits:
                fn = h.get("filename", "")
                if fn == "conversation":
                    label = "Past conversation"
                elif fn.startswith("_agent_ref_"):
                    label = "Industry reference material"
                elif fn.startswith("_"):
                    label = "Internal reference"
                else:
                    label = fn
                if fn != "conversation" and not fn.startswith("_"):
                    doc_filenames.add(fn)
                txt = (h.get("text") or "")[:1500]
                if txt:
                    rag_parts.append(f"\n### From {label}\n{txt}\n")
                    rag_sources.append({"filename": label, "text_preview": (txt or "")[:300]})
            if doc_filenames:
                rag_parts.append(f"\n\nIMPORTANT: You DO have access to the following documents: {', '.join(doc_filenames)}. "
                                 "Use the retrieved content above to answer the user's question. "
                                 "Do NOT say you don't have access to these documents.\n")
            _chat_log.info("RAG context: %d parts, total length %d", len(rag_parts), sum(len(p) for p in rag_parts))
        if is_graphrag_available():
            graphrag_texts = graphrag_search(request.message.strip(), top_k=3)
            if graphrag_texts:
                rag_parts.append("\n\n## GraphRAG context\n")
                rag_parts.append("\n".join(graphrag_texts[:3]))
        if rag_parts:
            system = system + "".join(rag_parts)
    except Exception:
        pass
    # Build message history: use client-provided history, or load persisted conversation so agent learns from all prior chat
    # Protection: filter out "poisoned" turns where the assistant incorrectly claimed
    # it lacked access to documents that are now available via RAG.
    _POISON_PHRASES = (
        "don't have access to",
        "do not have access to",
        "not available in my",
        "not in my reference library",
        "I cannot access",
        "I don't currently have",
        "outside my scope",
        "outside of my scope",
        "only help with scheduling",
        "only answer questions about project scheduling",
        "politely decline",
        "ask me a scheduling-related question",
    )

    def _is_poisoned(content: str) -> bool:
        cl = content.lower()
        return any(p in cl for p in _POISON_PHRASES)

    messages_for_llm = []
    for h in request.history:
        role = h.get("role")
        content = h.get("content") or ""
        if role in ("user", "assistant"):
            if role == "assistant" and _is_poisoned(content):
                if messages_for_llm and messages_for_llm[-1].get("role") == "user":
                    messages_for_llm.pop()
                continue
            messages_for_llm.append({"role": role, "content": content})
    if not messages_for_llm and request.session_id:
        from schedule_agent_web.store import get_conversation
        stored = get_conversation(request.session_id)
        for m in stored:
            r, c = m.get("role"), m.get("content") or ""
            if r in ("user", "assistant"):
                if r == "assistant" and _is_poisoned(c):
                    if messages_for_llm and messages_for_llm[-1].get("role") == "user":
                        messages_for_llm.pop()
                    continue
                messages_for_llm.append({"role": r, "content": c})
    # Keep only the last 20 turns (40 messages) to prevent stale context from dominating
    if len(messages_for_llm) > 40:
        messages_for_llm = messages_for_llm[-40:]
    messages_for_llm.append({"role": "user", "content": request.message.strip()})
    reply, err = _call_llm(system, messages_for_llm)
    if request.session_id and not err:
        try:
            from schedule_agent_web.store import append_to_conversation
            append_to_conversation(request.session_id, "user", request.message.strip())
            # Don't persist poisoned responses to the conversation store
            if not _is_poisoned(reply):
                append_to_conversation(request.session_id, "assistant", reply)
            else:
                append_to_conversation(request.session_id, "assistant",
                    "[System note: Previous answer was corrected — document content is now available via the reference library.]")
        except Exception:
            pass
        # Queue conversation for PCM digest review instead of auto-vectorizing.
        # FYI/Remember triggers get priority flagging.
        try:
            from schedule_agent_web.store import append_pending_digest
            msg_stripped = request.message.strip()
            cl = msg_stripped.lower()
            priority = "fyi" if (cl.startswith("fyi") or cl.startswith("remember this")) else "normal"
            turn_id = f"turn-{datetime.utcnow().isoformat()}-{hash(msg_stripped) & 0xFFFFFFFF:08x}"
            append_pending_digest(request.session_id, {
                "id": turn_id,
                "timestamp": datetime.utcnow().isoformat(),
                "user_message": msg_stripped,
                "assistant_response": reply,
                "priority": priority,
                "status": "pending",
                "topic": msg_stripped[:80] + ("..." if len(msg_stripped) > 80 else ""),
                "summary": "",
                "category": _guess_conversation_category(msg_stripped),
            })
        except Exception:
            pass
        # Stage 5: audit trail for chat answers
        if request.session_id and not err and reply:
            try:
                from schedule_agent_web.delivery import append_audit
                append_audit(request.session_id, "answer", {"source": "chat", "sources": rag_sources or []})
            except Exception:
                pass
    return ChatResponse(reply=reply, error=err, sources=rag_sources if rag_sources else None)


def handle_chat_json(message: str, history: list, session_id: str | None = None) -> dict:
    """Callable from api/chat.py: returns {"reply": str, "error": str|None}. Persists conversation if session_id and Redis set."""
    if not (message or "").strip():
        return {"reply": "", "error": "message is required"}
    if not get_status_dict().get("has_api_key"):
        return {"reply": "", "error": "Set OPENAI_API_KEY or ANTHROPIC_API_KEY in Vercel Environment Variables."}
    system = get_system_prompt_with_context(session_id)
    try:
        from schedule_agent_web.vector_store import search, graphrag_search, is_qdrant_available, is_graphrag_available, _file_search_fallback
        rag_parts = []
        if is_qdrant_available():
            hits = search(session_id, message.strip(), top_k=8)
        else:
            hits = _file_search_fallback(session_id, message.strip(), top_k=8) if session_id else []
        if hits:
            doc_filenames = set()
            rag_parts.append("\n\n## Retrieved from uploaded documents and past conversations (use when answering)\n")
            for h in hits:
                fn = h.get("filename", "")
                if fn == "conversation":
                    label = "Past conversation"
                elif fn.startswith("_agent_ref_"):
                    label = "Industry reference material"
                elif fn.startswith("_"):
                    label = "Internal reference"
                else:
                    label = fn
                if fn != "conversation" and not fn.startswith("_"):
                    doc_filenames.add(fn)
                txt = (h.get("text") or "")[:1500]
                if txt:
                    rag_parts.append(f"\n### From {label}\n{txt}\n")
            if doc_filenames:
                rag_parts.append(f"\n\nIMPORTANT: You DO have access to the following documents: {', '.join(doc_filenames)}. "
                                 "Use the retrieved content above to answer the user's question. "
                                 "Do NOT say you don't have access to these documents.\n")
        if is_graphrag_available():
            graphrag_texts = graphrag_search(message.strip(), top_k=3)
            if graphrag_texts:
                rag_parts.append("\n\n## GraphRAG context\n")
                rag_parts.append("\n".join(graphrag_texts[:3]))
        if rag_parts:
            system = system + "".join(rag_parts)
    except Exception:
        pass
    _POISON_PHRASES_ALT = (
        "don't have access to", "do not have access to",
        "not available in my", "not in my reference library",
        "I cannot access", "I don't currently have",
        "outside my scope", "outside of my scope",
        "only help with scheduling", "only answer questions about project scheduling",
        "politely decline", "ask me a scheduling-related question",
    )
    def _is_poisoned_alt(content: str) -> bool:
        cl = content.lower()
        return any(p in cl for p in _POISON_PHRASES_ALT)
    messages_for_llm = []
    for h in (history or []):
        role = h.get("role")
        content = h.get("content") or ""
        if role in ("user", "assistant"):
            if role == "assistant" and _is_poisoned_alt(content):
                if messages_for_llm and messages_for_llm[-1].get("role") == "user":
                    messages_for_llm.pop()
                continue
            messages_for_llm.append({"role": role, "content": content})
    if len(messages_for_llm) > 40:
        messages_for_llm = messages_for_llm[-40:]
    messages_for_llm.append({"role": "user", "content": message.strip()})
    reply, err = _call_llm(system, messages_for_llm)
    if session_id and not err:
        try:
            from schedule_agent_web.store import append_to_conversation
            append_to_conversation(session_id, "user", message.strip())
            if not _is_poisoned_alt(reply):
                append_to_conversation(session_id, "assistant", reply)
            else:
                append_to_conversation(session_id, "assistant",
                    "[System note: Previous answer was corrected — document content is now available via the reference library.]")
        except Exception:
            pass
        try:
            from schedule_agent_web.store import append_pending_digest
            msg_stripped = message.strip()
            cl = msg_stripped.lower()
            priority = "fyi" if (cl.startswith("fyi") or cl.startswith("remember this")) else "normal"
            turn_id = f"turn-{datetime.utcnow().isoformat()}-{hash(msg_stripped) & 0xFFFFFFFF:08x}"
            append_pending_digest(session_id, {
                "id": turn_id,
                "timestamp": datetime.utcnow().isoformat(),
                "user_message": msg_stripped,
                "assistant_response": reply,
                "priority": priority,
                "status": "pending",
                "topic": msg_stripped[:80] + ("..." if len(msg_stripped) > 80 else ""),
                "summary": "",
                "category": _guess_conversation_category(msg_stripped),
            })
        except Exception:
            pass
    return {"reply": reply, "error": err}


# Serve frontend locally (on Vercel, public/ is served by CDN)
if not os.environ.get("VERCEL"):
    static_dir = Path(__file__).resolve().parent / "static"
    if static_dir.exists():
        # Built VueLogic UI uses /assets/ for JS/CSS; serve them so the page isn't blank
        assets_dir = static_dir / "assets"
        if assets_dir.exists():
            app.mount("/assets", StaticFiles(directory=str(assets_dir)), name="assets")
        app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

        @app.get("/")
        def index():
            return FileResponse(
                static_dir / "index.html",
                headers={"Cache-Control": "no-cache, no-store, must-revalidate", "Pragma": "no-cache"},
            )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
