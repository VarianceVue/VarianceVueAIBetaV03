"""
Contractor Baseline Review — version-managed baseline submissions + AI review.

Each project (session_id) can have:
  - One set of contract specs (immutable once uploaded)
  - Multiple baseline submissions, each a (XER + optional narrative) pair
  - Submissions are auto-versioned: v1, v2, v3 …
  - Each submission can have one review result (Excel comments)

Schema lives in baseline_review.db alongside users.db.
"""
import os
import io
import json
import sqlite3
from datetime import datetime, timezone
from typing import Optional

_DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "baseline_review.db")


def _get_db():
    db = sqlite3.connect(_DB_PATH)
    db.row_factory = sqlite3.Row
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("""
        CREATE TABLE IF NOT EXISTS contract_specs (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  TEXT NOT NULL,
            filename    TEXT NOT NULL,
            size        INTEGER DEFAULT 0,
            uploaded_at TEXT NOT NULL,
            UNIQUE(session_id)
        )
    """)
    tables = [r[0] for r in db.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()]
    if "baseline_submissions" in tables:
        cols = [r[1] for r in db.execute("PRAGMA table_info(baseline_submissions)").fetchall()]
        if "submission_type" not in cols:
            db.execute("""
                CREATE TABLE baseline_submissions_new (
                    id              INTEGER PRIMARY KEY AUTOINCREMENT,
                    session_id      TEXT NOT NULL,
                    submission_type TEXT NOT NULL DEFAULT 'baseline',
                    version         INTEGER NOT NULL,
                    xer_filename    TEXT,
                    xer_size        INTEGER DEFAULT 0,
                    narr_filename   TEXT,
                    narr_size       INTEGER DEFAULT 0,
                    resp_filename   TEXT,
                    resp_size       INTEGER DEFAULT 0,
                    submitted_at    TEXT NOT NULL,
                    status          TEXT NOT NULL DEFAULT 'Under Review',
                    UNIQUE(session_id, submission_type, version)
                )
            """)
            src_cols = ", ".join(c for c in cols if c != "id")
            db.execute(f"""
                INSERT INTO baseline_submissions_new
                    (submission_type, {src_cols})
                SELECT 'baseline', {src_cols}
                FROM baseline_submissions
            """)
            db.execute("DROP TABLE baseline_submissions")
            db.execute("ALTER TABLE baseline_submissions_new RENAME TO baseline_submissions")
            db.commit()
    else:
        db.execute("""
            CREATE TABLE IF NOT EXISTS baseline_submissions (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                session_id      TEXT NOT NULL,
                submission_type TEXT NOT NULL DEFAULT 'baseline',
                version         INTEGER NOT NULL,
                xer_filename    TEXT,
                xer_size        INTEGER DEFAULT 0,
                narr_filename   TEXT,
                narr_size       INTEGER DEFAULT 0,
                resp_filename   TEXT,
                resp_size       INTEGER DEFAULT 0,
                submitted_at    TEXT NOT NULL,
                status          TEXT NOT NULL DEFAULT 'Under Review',
                UNIQUE(session_id, submission_type, version)
            )
        """)
    db.commit()
    return db


# ── Contract Specs ────────────────────────────────────────

def get_specs(session_id: str) -> Optional[dict]:
    db = _get_db()
    row = db.execute(
        "SELECT filename, size, uploaded_at FROM contract_specs WHERE session_id = ?",
        (session_id,),
    ).fetchone()
    db.close()
    if not row:
        return None
    return dict(row)


def save_specs_meta(session_id: str, filename: str, size: int) -> dict:
    db = _get_db()
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        "INSERT OR REPLACE INTO contract_specs (session_id, filename, size, uploaded_at) VALUES (?, ?, ?, ?)",
        (session_id, filename, size, now),
    )
    db.commit()
    db.close()
    return {"filename": filename, "size": size, "uploaded_at": now}


# ── Baseline Submissions ─────────────────────────────────

def _next_version(db, session_id: str, submission_type: str = "baseline") -> int:
    row = db.execute(
        "SELECT COALESCE(MAX(version), 0) AS mv FROM baseline_submissions "
        "WHERE session_id = ? AND submission_type = ?",
        (session_id, submission_type),
    ).fetchone()
    return (row["mv"] if row else 0) + 1


def list_submissions(session_id: str, submission_type: str | None = None) -> list[dict]:
    db = _get_db()
    if submission_type:
        rows = db.execute(
            "SELECT submission_type, version, xer_filename, xer_size, narr_filename, narr_size, "
            "resp_filename, resp_size, submitted_at, status "
            "FROM baseline_submissions WHERE session_id = ? AND submission_type = ? ORDER BY version DESC",
            (session_id, submission_type),
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT submission_type, version, xer_filename, xer_size, narr_filename, narr_size, "
            "resp_filename, resp_size, submitted_at, status "
            "FROM baseline_submissions WHERE session_id = ? ORDER BY submission_type, version DESC",
            (session_id,),
        ).fetchall()
    db.close()
    return [dict(r) for r in rows]


def next_expected_version(session_id: str, submission_type: str = "baseline") -> int:
    """Return the next version to show in the submission form.
    If the latest version for this type is incomplete (missing XER or narrative), return that version.
    Otherwise return latest+1."""
    db = _get_db()
    latest = db.execute(
        "SELECT version, xer_filename, narr_filename FROM baseline_submissions "
        "WHERE session_id = ? AND submission_type = ? ORDER BY version DESC LIMIT 1",
        (session_id, submission_type),
    ).fetchone()
    db.close()
    if not latest:
        return 1
    has_xer = bool(latest["xer_filename"])
    has_narr = bool(latest["narr_filename"])
    if has_xer and has_narr:
        return latest["version"] + 1
    return latest["version"]


def create_submission(
    session_id: str,
    submission_type: str = "baseline",
    xer_filename: str = "",
    xer_size: int = 0,
    narr_filename: Optional[str] = None,
    narr_size: int = 0,
    resp_filename: Optional[str] = None,
    resp_size: int = 0,
) -> dict:
    db = _get_db()
    ver = _next_version(db, session_id, submission_type)
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        "INSERT INTO baseline_submissions "
        "(session_id, submission_type, version, xer_filename, xer_size, narr_filename, narr_size, "
        "resp_filename, resp_size, submitted_at, status) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'Under Review')",
        (session_id, submission_type, ver, xer_filename, xer_size, narr_filename, narr_size,
         resp_filename, resp_size, now),
    )
    db.commit()
    db.close()
    return {
        "submission_type": submission_type,
        "version": ver,
        "xer_filename": xer_filename,
        "xer_size": xer_size,
        "narr_filename": narr_filename,
        "narr_size": narr_size,
        "resp_filename": resp_filename,
        "resp_size": resp_size,
        "submitted_at": now,
        "status": "Under Review",
    }


def get_submission(session_id: str, version: int, submission_type: str = "baseline") -> Optional[dict]:
    db = _get_db()
    row = db.execute(
        "SELECT submission_type, version, xer_filename, xer_size, narr_filename, narr_size, "
        "resp_filename, resp_size, submitted_at, status "
        "FROM baseline_submissions WHERE session_id = ? AND submission_type = ? AND version = ?",
        (session_id, submission_type, version),
    ).fetchone()
    db.close()
    return dict(row) if row else None


def update_submission(session_id: str, version: int, submission_type: str = "baseline", **kwargs) -> Optional[dict]:
    """Update specific fields on an existing submission."""
    db = _get_db()
    row = db.execute(
        "SELECT * FROM baseline_submissions WHERE session_id = ? AND submission_type = ? AND version = ?",
        (session_id, submission_type, version),
    ).fetchone()
    if not row:
        db.close()
        return None
    allowed = {"xer_filename", "xer_size", "narr_filename", "narr_size",
               "resp_filename", "resp_size", "status"}
    updates = {k: v for k, v in kwargs.items() if k in allowed and v is not None}
    if not updates:
        db.close()
        return dict(row)
    set_clause = ", ".join(f"{k} = ?" for k in updates)
    vals = list(updates.values()) + [session_id, submission_type, version]
    db.execute(
        f"UPDATE baseline_submissions SET {set_clause} "
        "WHERE session_id = ? AND submission_type = ? AND version = ?",
        vals,
    )
    db.commit()
    updated = db.execute(
        "SELECT submission_type, version, xer_filename, xer_size, narr_filename, narr_size, "
        "resp_filename, resp_size, submitted_at, status "
        "FROM baseline_submissions WHERE session_id = ? AND submission_type = ? AND version = ?",
        (session_id, submission_type, version),
    ).fetchone()
    db.close()
    return dict(updated) if updated else None


def get_previous_version(session_id: str, current_version: int) -> int:
    """Return the version number immediately before current_version, or 0 if none."""
    if current_version <= 1:
        return 0
    return current_version - 1


def delete_submissions_by_file(session_id: str, filename: str) -> int:
    """Delete all submissions that reference the given filename (XER, narrative, or response).
    Also deletes associated review results. Returns number of submissions removed."""
    db = _get_db()
    rows = db.execute(
        "SELECT submission_type, version FROM baseline_submissions WHERE session_id = ? "
        "AND (xer_filename = ? OR narr_filename = ? OR resp_filename = ?)",
        (session_id, filename, filename, filename),
    ).fetchall()
    count = 0
    for r in rows:
        db.execute(
            "DELETE FROM baseline_submissions WHERE session_id = ? AND submission_type = ? AND version = ?",
            (session_id, r["submission_type"], r["version"]),
        )
        count += 1
    db.commit()
    db.close()
    for r in rows:
        try:
            delete_review(session_id, r["version"])
        except Exception:
            pass
    return count


def has_scope_docs(session_id: str) -> bool:
    """Check if the project has any scope/ingestion documents uploaded (via store)."""
    try:
        from schedule_agent_web.store import get_files
        files = get_files(session_id)
        return len(files) > 0
    except Exception:
        return False


# ── Review Results ─────────────────────────────────────────

_REVIEWS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "file_store")

VERSIONED_COLS = {"Comment Description", "Recommendation for Correction", "Contractor Response"}

_COLUMN_TEMPLATE = [
    "Comment ID",
    "WBS Reference",
    "Comment Description",
    "Spec Section Reference",
    "Priority",
    "Logic Flag",
    "Recommendation for Correction",
    "Contractor Response",
    "Comment Status",
    "Addressed (Yes/No)",
]


def get_default_columns(version: int) -> list[str]:
    """Return default column names; version-specific fields get ' vN' suffix."""
    tag = f" v{version}"
    return [(c + tag if c in VERSIONED_COLS else c) for c in _COLUMN_TEMPLATE]


DEFAULT_COLUMNS = _COLUMN_TEMPLATE


def _reviews_dir(session_id: str) -> str:
    safe = "".join(c if c.isalnum() or c in "-_" else "_" for c in session_id)
    d = os.path.join(_REVIEWS_DIR, safe, "baseline_reviews")
    os.makedirs(d, exist_ok=True)
    return d


def save_review_result(session_id: str, version: int, comments: list[dict], columns: list[str],
                       source_files: dict | None = None) -> str:
    """Generate Excel workbook from comments, save to disk, return file path.

    source_files: optional dict with keys like 'xer', 'specs', 'narrative', 'response'
                  mapping to the original filenames used in the review.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Source Files sheet ---
    if source_files:
        ws_src = wb.active
        ws_src.title = "Review Source Files"
        title_font = Font(bold=True, size=14, color="1F4E79")
        label_font = Font(bold=True, size=11, color="333333")
        value_font = Font(size=11, color="0F172A")
        meta_fill = PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid")

        ws_src.merge_cells("A1:B1")
        title_cell = ws_src.cell(row=1, column=1, value="Review Input Files")
        title_cell.font = title_font

        row = 3
        label_map = [
            ("XER Schedule File", source_files.get("xer")),
            ("Contract Specifications", source_files.get("specs")),
            ("Schedule Narrative", source_files.get("narrative")),
            ("Contractor Response File", source_files.get("response")),
        ]
        for label, value in label_map:
            if value:
                lbl_cell = ws_src.cell(row=row, column=1, value=label)
                lbl_cell.font = label_font
                lbl_cell.fill = meta_fill
                val_cell = ws_src.cell(row=row, column=2, value=value)
                val_cell.font = value_font
                val_cell.fill = meta_fill
                row += 1

        ws_src.cell(row=row + 1, column=1, value="Review Generated").font = label_font
        ws_src.cell(row=row + 1, column=1).fill = meta_fill
        from datetime import datetime
        ws_src.cell(row=row + 1, column=2, value=datetime.now().strftime("%B %d, %Y at %I:%M %p")).font = value_font
        ws_src.cell(row=row + 1, column=2).fill = meta_fill

        ws_src.column_dimensions["A"].width = 30
        ws_src.column_dimensions["B"].width = 70

        ws = wb.create_sheet(title=f"Baseline v{version} Review")
    else:
        ws = wb.active
        ws.title = f"Baseline v{version} Review"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    open_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    logic_fills = {
        "Missing Hard Logic": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
        "Preferential Cross-WBS": PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid"),
        "Dangling Logic": PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
        "Redundant Logic": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
        "Incorrect Relationship Type": PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
    }
    logic_fonts = {
        "Missing Hard Logic": Font(bold=True, color="C62828"),
        "Preferential Cross-WBS": Font(bold=True, color="E65100"),
        "Dangling Logic": Font(bold=True, color="6A1B9A"),
        "Redundant Logic": Font(bold=True, color="1565C0"),
        "Incorrect Relationship Type": Font(bold=True, color="1565C0"),
    }
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, comment in enumerate(comments, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = comment.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border
            if col_name == "Comment Status" and val == "Open":
                cell.fill = open_fill
            if col_name == "Logic Flag" and val in logic_fills:
                cell.fill = logic_fills[val]
                cell.font = logic_fonts[val]
            base_col = col_name.rsplit(" v", 1)[0] if " v" in col_name else col_name
            if base_col == "Contractor Response":
                cell.value = ""
            if base_col == "Addressed (Yes/No)":
                cell.value = None

    _base_widths = {
        "Comment ID": 14, "WBS Reference": 22, "Comment Description": 50,
        "Spec Section Reference": 24, "Priority": 18, "Logic Flag": 24,
        "Recommendation for Correction": 50, "Contractor Response": 35,
        "Comment Status": 16, "Addressed (Yes/No)": 18,
    }
    from openpyxl.utils import get_column_letter
    for i, col_name in enumerate(columns, 1):
        base = col_name.rsplit(" v", 1)[0] if " v" in col_name else col_name
        ws.column_dimensions[get_column_letter(i)].width = _base_widths.get(base, 30)

    ws.auto_filter.ref = ws.dimensions

    out_dir = _reviews_dir(session_id)
    fname = f"Baseline_v{version}_Review_Comments.xlsx"
    path = os.path.join(out_dir, fname)
    wb.save(path)
    return path


_EXCEPTION_COLUMNS = [
    "Exception ID",
    "Original Comment ID",
    "WBS Reference",
    "Original Comment Description",
    "Spec Section Reference",
    "Contractor Response",
    "Contractor Claimed Addressed",
    "Agent Evaluation",
    "Evidence from XER",
    "Spec Non-Compliance Detail",
    "Severity",
]


def save_exception_report(session_id: str, version: int, exceptions: list[dict]) -> str:
    """Generate PCM Exception Report Excel from discrepancy findings."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"PCM Exception Report v{version}"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(_EXCEPTION_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    alert_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    wrap_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, exc in enumerate(exceptions, 2):
        for col_idx, col_name in enumerate(_EXCEPTION_COLUMNS, 1):
            val = exc.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap_align
            cell.border = thin_border
            if col_name == "Agent Evaluation":
                cell.fill = alert_fill
                cell.font = Font(bold=True, color="8B0000")

    exc_widths = {
        "Exception ID": 14, "Original Comment ID": 18, "WBS Reference": 22,
        "Original Comment Description": 50, "Spec Section Reference": 24,
        "Contractor Response": 40, "Contractor Claimed Addressed": 20,
        "Agent Evaluation": 18, "Evidence from XER": 50,
        "Spec Non-Compliance Detail": 50, "Severity": 14,
    }
    for i, col_name in enumerate(_EXCEPTION_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = exc_widths.get(col_name, 30)

    ws.auto_filter.ref = ws.dimensions

    out_dir = _reviews_dir(session_id)
    fname = f"PCM_Exception_Report_v{version}.xlsx"
    path = os.path.join(out_dir, fname)
    wb.save(path)
    return path


def list_reviews(session_id: str) -> list[dict]:
    """Return metadata for all review results."""
    db = _get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS baseline_reviews (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  TEXT NOT NULL,
            version     INTEGER NOT NULL,
            columns     TEXT NOT NULL,
            comment_count INTEGER DEFAULT 0,
            last_comment_num INTEGER DEFAULT 0,
            exception_count INTEGER DEFAULT 0,
            exception_filepath TEXT DEFAULT '',
            filepath    TEXT NOT NULL,
            created_at  TEXT NOT NULL,
            UNIQUE(session_id, version)
        )
    """)
    for col_sql in [
        "ALTER TABLE baseline_reviews ADD COLUMN last_comment_num INTEGER DEFAULT 0",
        "ALTER TABLE baseline_reviews ADD COLUMN exception_count INTEGER DEFAULT 0",
        "ALTER TABLE baseline_reviews ADD COLUMN exception_filepath TEXT DEFAULT ''",
    ]:
        try:
            db.execute(col_sql)
        except Exception:
            pass
    db.commit()
    rows = db.execute(
        "SELECT version, columns, comment_count, exception_count, exception_filepath, filepath, created_at "
        "FROM baseline_reviews WHERE session_id = ? ORDER BY version DESC",
        (session_id,),
    ).fetchall()
    db.close()
    results = []
    for r in rows:
        d = dict(r)
        d["has_exception_report"] = bool(d.get("exception_filepath"))
        results.append(d)
    return results


def get_last_comment_number(session_id: str) -> int:
    """Return the highest BLR-NNN number used across all reviews for this project."""
    db = _get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS baseline_reviews (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  TEXT NOT NULL,
            version     INTEGER NOT NULL,
            columns     TEXT NOT NULL,
            comment_count INTEGER DEFAULT 0,
            last_comment_num INTEGER DEFAULT 0,
            filepath    TEXT NOT NULL,
            created_at  TEXT NOT NULL,
            UNIQUE(session_id, version)
        )
    """)
    db.commit()
    try:
        db.execute("ALTER TABLE baseline_reviews ADD COLUMN last_comment_num INTEGER DEFAULT 0")
        db.commit()
    except Exception:
        pass
    row = db.execute(
        "SELECT COALESCE(MAX(last_comment_num), 0) AS mx FROM baseline_reviews WHERE session_id = ?",
        (session_id,),
    ).fetchone()
    db.close()
    return row["mx"] if row else 0


def save_review_meta(
    session_id: str, version: int, columns: list[str], comment_count: int,
    filepath: str, last_comment_num: int = 0,
    exception_count: int = 0, exception_filepath: str = "",
) -> dict:
    db = _get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS baseline_reviews (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  TEXT NOT NULL,
            version     INTEGER NOT NULL,
            columns     TEXT NOT NULL,
            comment_count INTEGER DEFAULT 0,
            last_comment_num INTEGER DEFAULT 0,
            exception_count INTEGER DEFAULT 0,
            exception_filepath TEXT DEFAULT '',
            filepath    TEXT NOT NULL,
            created_at  TEXT NOT NULL,
            UNIQUE(session_id, version)
        )
    """)
    for col_sql in [
        "ALTER TABLE baseline_reviews ADD COLUMN last_comment_num INTEGER DEFAULT 0",
        "ALTER TABLE baseline_reviews ADD COLUMN exception_count INTEGER DEFAULT 0",
        "ALTER TABLE baseline_reviews ADD COLUMN exception_filepath TEXT DEFAULT ''",
    ]:
        try:
            db.execute(col_sql)
        except Exception:
            pass
    now = datetime.now(timezone.utc).isoformat()
    db.execute(
        "INSERT OR REPLACE INTO baseline_reviews "
        "(session_id, version, columns, comment_count, last_comment_num, "
        "exception_count, exception_filepath, filepath, created_at) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (session_id, version, json.dumps(columns), comment_count, last_comment_num,
         exception_count, exception_filepath, filepath, now),
    )
    db.commit()
    db.close()
    return {
        "version": version, "columns": columns, "comment_count": comment_count,
        "exception_count": exception_count, "has_exception_report": bool(exception_filepath),
        "filepath": filepath, "created_at": now,
    }


def delete_review(session_id: str, version: int) -> bool:
    """Delete a review result (DB row + files on disk + vector store). Returns True if deleted."""
    db = _get_db()
    row = db.execute(
        "SELECT filepath, exception_filepath FROM baseline_reviews "
        "WHERE session_id = ? AND version = ?",
        (session_id, version),
    ).fetchone()
    if not row:
        db.close()
        return False
    for fpath in [row["filepath"], row["exception_filepath"]]:
        if fpath and os.path.isfile(fpath):
            try:
                os.remove(fpath)
            except OSError:
                pass
    db.execute(
        "DELETE FROM baseline_reviews WHERE session_id = ? AND version = ?",
        (session_id, version),
    )
    db.commit()
    db.close()
    try:
        from schedule_agent_web.vector_store import delete_file_vectors
        review_vec_name = f"_baseline_review_v{version}"
        delete_file_vectors(session_id, review_vec_name)
        exception_vec_name = f"_baseline_exception_v{version}"
        delete_file_vectors(session_id, exception_vec_name)
    except Exception:
        pass
    return True


def get_review(session_id: str, version: int) -> Optional[dict]:
    db = _get_db()
    db.execute("""
        CREATE TABLE IF NOT EXISTS baseline_reviews (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            session_id  TEXT NOT NULL,
            version     INTEGER NOT NULL,
            columns     TEXT NOT NULL,
            comment_count INTEGER DEFAULT 0,
            last_comment_num INTEGER DEFAULT 0,
            exception_count INTEGER DEFAULT 0,
            exception_filepath TEXT DEFAULT '',
            filepath    TEXT NOT NULL,
            created_at  TEXT NOT NULL,
            UNIQUE(session_id, version)
        )
    """)
    for col_sql in [
        "ALTER TABLE baseline_reviews ADD COLUMN last_comment_num INTEGER DEFAULT 0",
        "ALTER TABLE baseline_reviews ADD COLUMN exception_count INTEGER DEFAULT 0",
        "ALTER TABLE baseline_reviews ADD COLUMN exception_filepath TEXT DEFAULT ''",
    ]:
        try:
            db.execute(col_sql)
        except Exception:
            pass
    db.commit()
    row = db.execute(
        "SELECT version, columns, comment_count, exception_count, exception_filepath, filepath, created_at "
        "FROM baseline_reviews WHERE session_id = ? AND version = ?",
        (session_id, version),
    ).fetchone()
    db.close()
    if not row:
        return None
    d = dict(row)
    d["columns"] = json.loads(d["columns"]) if isinstance(d["columns"], str) else d["columns"]
    d["has_exception_report"] = bool(d.get("exception_filepath"))
    return d
